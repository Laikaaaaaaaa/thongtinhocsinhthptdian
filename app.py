from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, after_this_request
from flask_cors import CORS
import sqlite3
import pandas as pd
import json
from datetime import datetime, timedelta
import os
import math
import smtplib
import random
import time
import uuid
import string
import hashlib
import glob
from urllib.parse import quote
import urllib.parse
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv

# Try to import PostgreSQL support for Heroku
try:
    import psycopg2
    POSTGRES_AVAILABLE = True
except ImportError:
    POSTGRES_AVAILABLE = False

env_path = os.path.join(os.path.dirname(__file__), 'configs', 'security', 'environment', 'production', 'secrets', 'app', 'database', 'email', 'admin', 'settings', '.env')
load_dotenv(env_path)

def get_placeholder():
    """Get the correct placeholder for current database type"""
    return '%s' if DB_CONFIG['type'] == 'postgresql' else '?'

def convert_placeholders(query, use_postgres=None):
    """Convert query placeholders based on database type"""
    if use_postgres is None:
        use_postgres = (DB_CONFIG['type'] == 'postgresql')
    
    if use_postgres:
        return query.replace('?', '%s')
    else:
        return query.replace('%s', '?')

app = Flask(__name__)
CORS(app)
app.secret_key = os.getenv('SECRET_KEY', 'thpt-di-an-secret-key-2025')

EMAIL_CONFIG = {
    'smtp_server': os.getenv('SMTP_SERVER', 'smtp.gmail.com'),
    'smtp_port': int(os.getenv('SMTP_PORT', 587)),
    'email': os.getenv('SMTP_EMAIL', 'your-email@gmail.com'),
    'password': os.getenv('SMTP_PASSWORD', 'your-app-password'),
    'timeout': 30
}

# Database Configuration - Support both SQLite (local) and PostgreSQL (Heroku)
DATABASE_URL = os.getenv('DATABASE_URL')
if DATABASE_URL and POSTGRES_AVAILABLE:
    # Parse Heroku PostgreSQL URL
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://')
    
    url = urllib.parse.urlparse(DATABASE_URL)
    DB_CONFIG = {
        'type': 'postgresql',
        'host': url.hostname,
        'port': url.port,
        'database': url.path[1:],  # Remove leading slash
        'user': url.username,
        'password': url.password
    }
    print("🐘 Using PostgreSQL database (Heroku) - v2")
else:
    DB_CONFIG = {
        'type': 'sqlite',
        'path': 'students.db'
    }
    print("📁 Using SQLite database (local)")

def get_db_connection():
    """Get database connection based on configuration"""
    if DB_CONFIG['type'] == 'postgresql':
        import psycopg2
        return psycopg2.connect(
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port'],
            database=DB_CONFIG['database'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password']
        )
    else:
        return sqlite3.connect(DB_CONFIG['path'])

def init_database():
    """Initialize database with students table"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if DB_CONFIG['type'] == 'postgresql':
            # PostgreSQL syntax
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS students (
                    id SERIAL PRIMARY KEY,
                    ho_ten VARCHAR(255),
                    ngay_sinh DATE,
                    gioi_tinh VARCHAR(10),
                    lop VARCHAR(20),
                    khoi VARCHAR(10),
                    sdt VARCHAR(20),
                    email VARCHAR(255),
                    dia_chi TEXT,
                    tinh_thanh VARCHAR(100),
                    dan_toc VARCHAR(50),
                    ton_giao VARCHAR(50),
                    ho_ten_cha VARCHAR(255),
                    nghe_nghiep_cha VARCHAR(100),
                    ho_ten_me VARCHAR(255),
                    nghe_nghiep_me VARCHAR(100),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        else:
            # SQLite syntax
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS students (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ho_ten TEXT,
                    ngay_sinh DATE,
                    gioi_tinh TEXT,
                    lop TEXT,
                    khoi TEXT,
                    sdt TEXT,
                    email TEXT,
                    dia_chi TEXT,
                    tinh_thanh TEXT,
                    dan_toc TEXT,
                    ton_giao TEXT,
                    ho_ten_cha TEXT,
                    nghe_nghiep_cha TEXT,
                    ho_ten_me TEXT,
                    nghe_nghiep_me TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        
        conn.commit()
        conn.close()
        print(f"[DB] ✅ Database initialized with {DB_CONFIG['type']}")
        
    except Exception as e:
        print(f"[DB] ❌ Database initialization failed: {e}")

# Initialize database on startup
init_database()

otp_storage = {}

def parse_admin_accounts():
    """Parse admin accounts from environment variable"""
    admin_str = os.getenv('ADMIN_ACCOUNTS', 'admin@example.com:password123')
    accounts = {}
    try:
        for pair in admin_str.split(','):
            email, password = pair.split(':')
            accounts[email.strip()] = password.strip()
    except Exception as e:
        print(f"[CONFIG] Error parsing ADMIN_ACCOUNTS: {e}")
        accounts = {'admin@example.com': 'password123'}
    return accounts

ADMIN_ACCOUNTS = parse_admin_accounts()

def cleanup_file(filepath):
    """Helper function to delete a file safely with retry mechanism"""
    import threading
    import time
    
    def delayed_cleanup():
        max_retries = 5
        retry_delay = 2  # seconds
        
        for attempt in range(max_retries):
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    print(f"[CLEANUP] ✅ Deleted file: {os.path.basename(filepath)} (attempt {attempt + 1})")
                    return
                else:
                    print(f"[CLEANUP] ℹ️ File already removed: {os.path.basename(filepath)}")
                    return
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"[CLEANUP] ⏳ Retry {attempt + 1}/{max_retries} for {os.path.basename(filepath)} in {retry_delay}s: {e}")
                    time.sleep(retry_delay)
                    retry_delay *= 2  # Exponential backoff
                else:
                    print(f"[CLEANUP] ❌ Failed to delete {filepath} after {max_retries} attempts: {e}")
    
    # Schedule cleanup in background thread with delay
    cleanup_thread = threading.Thread(target=delayed_cleanup)
    cleanup_thread.daemon = True
    cleanup_thread.start()

def send_file_with_cleanup(filepath, **kwargs):
    """Send file and schedule it for deletion after download"""
    @after_this_request
    def remove_file(response):
        cleanup_file(filepath)
        return response
    
    # Handle UTF-8 encoding for Vietnamese filenames
    if 'download_name' in kwargs:
        download_name = kwargs['download_name']
        # Remove download_name from kwargs to avoid duplicate parameter
        kwargs_copy = kwargs.copy()
        del kwargs_copy['download_name']
        
        # Send file and manually set Content-Disposition header with UTF-8 encoding
        response = send_file(filepath, **kwargs_copy)
        encoded_name = quote(download_name.encode('utf-8'), safe='')
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_name}'
        return response
    
    return send_file(filepath, **kwargs)

DEBUG_OTP = os.getenv('DEBUG_OTP', 'true').lower() == 'true'
FORCE_CONSOLE_OTP = os.getenv('FORCE_CONSOLE_OTP', 'false').lower() == 'true'
SHOW_ADMIN_CREDENTIALS = os.getenv('SHOW_ADMIN_CREDENTIALS', 'false').lower() == 'true'
@app.before_request
def _log_request():
    try:
        print(f"[REQ] {request.method} {request.path}")
    except Exception:
        pass
def test_email_config():
    """Test email configuration at startup"""
    try:
        print(f"[EMAIL CONFIG] Testing connection to {EMAIL_CONFIG['smtp_server']}...")
        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'], timeout=EMAIL_CONFIG['timeout'])
        server.starttls()
        server.login(EMAIL_CONFIG['email'], EMAIL_CONFIG['password'])
        server.quit()
        print(f"[EMAIL CONFIG] ✅ Email configuration is working!")
        return True
    except Exception as e:
        print(f"[EMAIL CONFIG] ❌ Email configuration failed: {str(e)}")
        print(f"[EMAIL CONFIG] OTP will be shown in console only")
        return False
def send_otp_email(email, otp):
    """Gửi OTP qua email với fallback console logging"""

    if DEBUG_OTP:
        print(f"\n" + "="*50)
        print(f"🔐 OTP ADMIN LOGIN - DEBUG MODE")
        print(f"📧 Email: {email}")
        print(f"🔢 OTP Code: {otp}")
        print(f"⏰ Valid for 5 minutes")
        print(f"="*50 + "\n")

    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = EMAIL_CONFIG['email']
        msg['To'] = email
        msg['Subject'] = 'THPT Di An - Ma OTP Admin'

        text_body = f"""
THPT Di An - He thong quan ly hoc sinh

Ma xac thuc dang nhap Admin: {otp}

Ma co hieu luc trong 5 phut.
Khong chia se ma nay voi bat ky ai khac.

© 2025 THPT Di An
        """

        html_body = f"""
        <html>
        <head><meta charset="UTF-8"></head>
        <body style="font-family: Arial, sans-serif; margin: 20px; color: #333;">
            <div style="max-width: 500px; margin: 0 auto;">
                <h2 style="color: #2196F3;">🏫 THPT Dĩ An</h2>
                <p>Hệ thống quản lý học sinh</p>

                <div style="background: #f0f8ff; padding: 20px; border-radius: 5px; text-align: center; margin: 20px 0;">
                    <h3 style="margin: 0 0 10px 0; color: #1565C0;">Mã xác thực Admin</h3>
                    <div style="font-size: 28px; font-weight: bold; color: #2196F3; letter-spacing: 3px;">{otp}</div>
                    <p style="margin: 10px 0 0 0; font-size: 14px; color: #666;">Có hiệu lực trong 5 phút</p>
                </div>

                <p style="font-size: 13px; color: #666;">
                    ⚠️ Không chia sẻ mã này với bất kỳ ai khác.<br>
                    © 2025 THPT Dĩ An - Hệ thống quản lý học sinh
                </p>
            </div>
        </body>
        </html>
        """

        text_part = MIMEText(text_body, 'plain', 'utf-8')
        html_part = MIMEText(html_body, 'html', 'utf-8')

        msg.attach(text_part)
        msg.attach(html_part)

        server = None
        try:
            print(f"[EMAIL] Attempting to send OTP to {email}...")
            server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'], timeout=EMAIL_CONFIG['timeout'])
            server.set_debuglevel(0)
            server.starttls()
            server.login(EMAIL_CONFIG['email'], EMAIL_CONFIG['password'])

            result = server.send_message(msg)
            print(f"[EMAIL] ✅ OTP sent successfully to {email}")
            return True

        except smtplib.SMTPAuthenticationError as e:
            print(f"[EMAIL] ❌ Authentication failed: {str(e)}")
            print(f"[EMAIL] Check email credentials and app password")
            return False

        except smtplib.SMTPRecipientsRefused as e:
            print(f"[EMAIL] ❌ Recipients refused: {str(e)}")
            return False

        except smtplib.SMTPServerDisconnected as e:
            print(f"[EMAIL] ❌ Server disconnected: {str(e)}")
            return False

        except Exception as e:
            print(f"[EMAIL] ❌ Unexpected error: {str(e)}")
            print(f"[EMAIL] Error type: {type(e).__name__}")
            return False

        finally:
            if server:
                try:
                    server.quit()
                except:
                    pass

    except Exception as e:
        print(f"[EMAIL] ❌ Failed to create email: {str(e)}")
        return False

def generate_otp():
    """Tạo OTP 6 số"""
    return str(random.randint(100000, 999999))

def store_otp(email, otp):
    """Lưu OTP với thời gian hết hạn"""
    otp_storage[email] = {
        'otp': otp,
        'created_at': time.time(),
        'expires_at': time.time() + 300
    }

def verify_otp(email, otp):
    """Xác thực OTP"""
    if email not in otp_storage:
        return False, "Không tìm thấy mã OTP"

    stored_data = otp_storage[email]
    current_time = time.time()

    if current_time > stored_data['expires_at']:
        del otp_storage[email]
        return False, "Mã OTP đã hết hạn"

    if stored_data['otp'] != otp:
        return False, "Mã OTP không đúng"

    del otp_storage[email]
    return True, "Xác thực thành công"

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('PRAGMA journal_mode = WAL')
    cursor.execute('PRAGMA synchronous = NORMAL')
    cursor.execute('PRAGMA cache_size = 10000')
    cursor.execute('PRAGMA temp_store = MEMORY')
    cursor.execute('PRAGMA mmap_size = 268435456')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL,
            full_name TEXT,
            nickname TEXT,
            class TEXT,
            birth_date TEXT,
            gender TEXT,
            ethnicity TEXT,
            nationality TEXT,
            religion TEXT,
            phone TEXT,
            citizen_id TEXT,
            cccd_date TEXT,
            cccd_place TEXT,
            personal_id TEXT,
            passport TEXT,
            passport_date TEXT,
            passport_place TEXT,
            organization TEXT,
            permanent_province TEXT,
            permanent_ward TEXT,
            permanent_hamlet TEXT,
            permanent_street TEXT,
            hometown_province TEXT,
            hometown_ward TEXT,
            hometown_hamlet TEXT,
            birth_cert_province TEXT,
            birth_cert_ward TEXT,
            birthplace_province TEXT,
            birthplace_ward TEXT,
            current_address_detail TEXT,
            current_province TEXT,
            current_ward TEXT,
            current_hamlet TEXT,
            height REAL,
            weight REAL,
            eye_diseases TEXT,
            swimming_skill TEXT,
            smartphone TEXT,
            computer TEXT,
            father_ethnicity TEXT,
            mother_ethnicity TEXT,
            father_name TEXT,
            father_job TEXT,
            father_birth_year TEXT,
            father_phone TEXT,
            father_cccd TEXT,
            mother_name TEXT,
            mother_job TEXT,
            mother_birth_year TEXT,
            mother_phone TEXT,
            mother_cccd TEXT,
            guardian_name TEXT,
            guardian_job TEXT,
            guardian_birth_year TEXT,
            guardian_phone TEXT,
            guardian_cccd TEXT,
            guardian_gender TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('CREATE INDEX IF NOT EXISTS idx_email ON students(email)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_created_at ON students(created_at)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_full_name ON students(full_name)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_class ON students(class)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_id_email ON students(id, email)')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS otp_codes_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL,
            otp_code TEXT NOT NULL,
            purpose TEXT NOT NULL DEFAULT 'admin_login',
            expires_at TIMESTAMP NOT NULL,
            used BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    try:
        cursor.execute('SELECT * FROM otp_codes LIMIT 1')
        if cursor.fetchone():
            cursor.execute('''
                INSERT INTO otp_codes_new (email, otp_code, purpose, expires_at, used, created_at)
                SELECT email, otp_code, 'admin_login', expires_at, used, created_at FROM otp_codes
            ''')
            cursor.execute('DROP TABLE otp_codes')
    except sqlite3.OperationalError:
        pass

    cursor.execute('DROP TABLE IF EXISTS otp_codes')
    cursor.execute('ALTER TABLE otp_codes_new RENAME TO otp_codes')

    cursor.execute('CREATE INDEX IF NOT EXISTS idx_otp_email_purpose ON otp_codes(email, purpose)')

    conn.commit()
    conn.close()
    print("[DB] ✅ Database initialized with performance optimizations for 1000+ records")

def migrate_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('PRAGMA table_info(students)')
        existing_cols = {row[1] for row in cursor.fetchall()}

        expected_cols = [
            ('height', 'REAL'),
            ('weight', 'REAL'),
            ('eye_diseases', 'TEXT'),
            ('swimming_skill', 'TEXT'),
            ('smartphone', 'TEXT'),
            ('computer', 'TEXT'),
            ('father_ethnicity', 'TEXT'),
            ('mother_ethnicity', 'TEXT'),
            ('father_name', 'TEXT'),
            ('father_job', 'TEXT'),
            ('father_birth_year', 'TEXT'),
            ('father_phone', 'TEXT'),
            ('father_cccd', 'TEXT'),
            ('mother_name', 'TEXT'),
            ('mother_job', 'TEXT'),
            ('mother_birth_year', 'TEXT'),
            ('mother_phone', 'TEXT'),
            ('mother_cccd', 'TEXT'),
            ('guardian_name', 'TEXT'),
            ('guardian_job', 'TEXT'),
            ('guardian_birth_year', 'TEXT'),
            ('guardian_phone', 'TEXT'),
            ('guardian_cccd', 'TEXT'),
            ('guardian_gender', 'TEXT'),
            ('created_at', 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
        ]

        for col, col_type in expected_cols:
            if col not in existing_cols:
                cursor.execute(f'ALTER TABLE students ADD COLUMN {col} {col_type}')

        conn.commit()
    finally:
        conn.close()

def generate_sample_data(count=150):
    """Tạo dữ liệu mẫu cho testing"""
    import random
    from datetime import datetime, timedelta

    first_names = [
        'Nguyễn', 'Trần', 'Lê', 'Phạm', 'Hoàng', 'Huỳnh', 'Phan', 'Vũ', 'Võ', 'Đặng',
        'Bùi', 'Đỗ', 'Hồ', 'Ngô', 'Dương', 'Lý', 'Đinh', 'Đào', 'Cao', 'Lương'
    ]

    middle_names = ['Văn', 'Thị', 'Minh', 'Hoàng', 'Quang', 'Hữu', 'Thanh', 'Anh', 'Thành', 'Bảo']

    last_names = [
        'An', 'Bình', 'Cường', 'Dũng', 'Đức', 'Giang', 'Hà', 'Hải', 'Khang', 'Linh',
        'Long', 'Mai', 'Minh', 'Nam', 'Phong', 'Quân', 'Sơn', 'Thảo', 'Tú', 'Vy',
        'Yến', 'Hương', 'Loan', 'Nga', 'Oanh', 'Phương', 'Quyên', 'Thu', 'Trang', 'Xuân'
    ]

    classes = [
        # Khối 10
        '10A1', '10A2', '10A3', '10A4', '10A5', '10A6', '10A7', '10A8',
        '10B1', '10B2', '10B3', '10B4',
        # Khối 11
        '11A1', '11A2', '11A3', '11A4', '11A5', '11A6', '11A7', '11A8',
        '11B1', '11B2', '11B3', '11B4',
        # Khối 12
        '12A1', '12A2', '12A3', '12A4', '12A5', '12A6', '12A7', '12A8',
        '12B1', '12B2', '12B3', '12B4'
    ]

    conducts = ['Tốt', 'Khá', 'Trung bình']
    performances = ['Giỏi', 'Khá', 'Trung bình']
    genders = ['Nam', 'Nữ']

    provinces = ['Đồng Nai', 'TP.HCM', 'Bình Dương', 'Long An', 'Tây Ninh']

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        print(f"[SAMPLE DATA] Tạo {count} bản ghi mẫu...")

        for i in range(count):
            first_name = random.choice(first_names)
            middle_name = random.choice(middle_names)
            last_name = random.choice(last_names)
            full_name = f"{first_name} {middle_name} {last_name}"

            email_name = f"{last_name.lower()}.{middle_name.lower()}.{i+100:03d}"
            email = f"{email_name}@student.dian.edu.vn"

            birth_year = random.randint(2005, 2008)
            birth_month = random.randint(1, 12)
            birth_day = random.randint(1, 28)
            birth_date = f"{birth_day:02d}/{birth_month:02d}/{birth_year}"

            gender = random.choice(genders)
            phone = f"0{random.randint(700000000, 999999999)}"
            class_name = random.choice(classes)
            gpa = round(random.uniform(6.5, 9.5), 1)
            conduct = random.choice(conducts)
            performance = random.choice(performances)
            province = random.choice(provinces)

            street_num = random.randint(1, 500)
            current_address = f"{street_num} Đường {random.randint(1, 50)}"

            days_ago = random.randint(0, 30)
            created_at = datetime.now() - timedelta(days=days_ago, hours=random.randint(0, 23), minutes=random.randint(0, 59))

            cursor.execute('''
                INSERT INTO students (
                    email, full_name, birth_date, gender, phone, current_address_detail,
                    class, current_province, created_at,
                    birthplace_province, birthplace_ward,
                    height, weight, smartphone, computer, nationality, ethnicity
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                email, full_name, birth_date, gender, phone, current_address,
                class_name, province, created_at.isoformat(),
                province, f"Phường {random.randint(1, 20)}",
                random.randint(150, 180), random.randint(45, 75),
                random.choice(['Có', 'Không']), random.choice(['Có', 'Không']),
                'Việt Nam', 'Kinh'
            ))

            if (i + 1) % 50 == 0:
                print(f"[SAMPLE DATA] Đã tạo {i + 1}/{count} bản ghi...")

        conn.commit()
        print(f"[SAMPLE DATA] ✅ Đã tạo thành công {count} bản ghi mẫu!")

    except Exception as e:
        print(f"[SAMPLE DATA] ❌ Lỗi tạo dữ liệu mẫu: {e}")
        conn.rollback()
    finally:
        conn.close()

@app.route('/')
def index():
    return send_file('page.html')

# Favicon routes
@app.route('/favicon.ico')
def favicon():
    response = app.response_class(
        response=open('logo/favicon.ico', 'rb').read(),
        status=200,
        mimetype='image/x-icon'
    )
    # No cache để force reload favicon mới
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/logo/<path:filename>')
def serve_logo(filename):
    response = send_file(f'logo/{filename}')
    # No cache để force reload logo mới
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache' 
    response.headers['Expires'] = '0'
    return response

@app.route('/apple-touch-icon.png')
def apple_touch_icon():
    response = send_file('logo/apple-touch-icon.png')
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/android-chrome-192x192.png')
def android_chrome_192():
    response = send_file('logo/android-chrome-192x192.png')
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/android-chrome-512x512.png')
def android_chrome_512():
    response = send_file('logo/android-chrome-512x512.png')
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# Test favicon page để kiểm tra favicon với timestamp
@app.route('/test-favicon.html')
def test_favicon():
    import time
    timestamp = int(time.time())
    html_content = f'''
<!DOCTYPE html>
<html>
<head>
    <title>Test Favicon - THPT Dĩ An</title>
    <!-- Force reload favicon với timestamp -->
    <link rel="icon" type="image/x-icon" href="/favicon.ico?v={timestamp}">
    <link rel="icon" type="image/png" sizes="32x32" href="/logo/favicon-32x32.png?v={timestamp}">
    <link rel="icon" type="image/png" sizes="16x16" href="/logo/favicon-16x16.png?v={timestamp}">
    <link rel="apple-touch-icon" sizes="180x180" href="/logo/apple-touch-icon.png?v={timestamp}">
    <link rel="icon" type="image/png" sizes="192x192" href="/logo/android-chrome-192x192.png?v={timestamp}">
    <link rel="icon" type="image/png" sizes="512x512" href="/logo/android-chrome-512x512.png?v={timestamp}">
    <link rel="manifest" href="/logo/site.webmanifest?v={timestamp}">
    <style>
        body {{ font-family: Arial, sans-serif; padding: 20px; }}
        .favicon-test {{ border: 2px solid #ddd; padding: 20px; margin: 10px 0; }}
        .timestamp {{ color: #666; font-size: 12px; }}
    </style>
</head>
<body>
    <h1>🔍 Test Favicon - THPT Dĩ An</h1>
    <div class="favicon-test">
        <h2>✅ Favicon đã được tải với timestamp: {timestamp}</h2>
        <p>Nếu bạn thấy favicon cũ, hãy:</p>
        <ul>
            <li>🔄 Nhấn <strong>Ctrl + F5</strong> để hard refresh</li>
            <li>🗑️ Xóa cache trình duyệt</li>
            <li>📱 Mở tab mới hoặc cửa sổ ẩn danh</li>
        </ul>
        <p class="timestamp">Timestamp: {timestamp}</p>
        <p>Favicon URLs với cache busting:</p>
        <ul>
            <li><a href="/favicon.ico?v={timestamp}" target="_blank">favicon.ico</a></li>
            <li><a href="/logo/favicon-32x32.png?v={timestamp}" target="_blank">favicon-32x32.png</a></li>
            <li><a href="/logo/apple-touch-icon.png?v={timestamp}" target="_blank">apple-touch-icon.png</a></li>
        </ul>
    </div>
</body>
</html>
    '''
    response = app.response_class(
        response=html_content,
        status=200,
        mimetype='text/html'
    )
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/page.html')
def page():
    return send_file('page.html')

@app.route('/page1.html')
def page1():
    return send_file('page1.html')

@app.route('/page2.html')
def page2():
    return send_file('page2.html')

@app.route('/page3.html')
def page3():
    return send_file('page3.html')

@app.route('/page4.html')
def page4():
    return send_file('page4.html')

@app.route('/page5.html')
def page5():
    return send_file('page5.html')

@app.route('/done.html')
def done():
    return send_file('done.html')

@app.route('/api/save-student', methods=['POST', 'OPTIONS'])
@app.route('/api/save-student/', methods=['POST', 'OPTIONS'])
def save_student():
    if request.method == 'OPTIONS':
        return ('', 200)
    try:
        data = request.get_json()
        if not data or not data.get('email'):
            return jsonify({'success': False, 'message': 'Thiếu email đăng ký'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        if DB_CONFIG['type'] == 'postgresql':
            cursor.execute('SELECT id FROM students WHERE email = %s', (data.get('email'),))
        else:
            cursor.execute('SELECT id FROM students WHERE email = ?', (data.get('email'),))
        existing = cursor.fetchone()

        col_map = [
            ('email', 'email'),
            ('ho_ten', 'fullName'),
            ('lop', 'class'),
            ('khoi', 'grade'),
            ('ngay_sinh', 'birthDate'),
            ('gioi_tinh', 'gender'),
            ('sdt', 'phone'),
            ('dan_toc', 'ethnicity'),
            ('ton_giao', 'religion'),
            ('dia_chi', 'currentAddressDetail'),
            ('tinh_thanh', 'currentProvince'),
            ('ho_ten_cha', 'fatherName'),
            ('nghe_nghiep_cha', 'fatherJob'),
            ('ho_ten_me', 'motherName'),
            ('nghe_nghiep_me', 'motherJob'),
            # Personal info extended
            ('nickname', 'nickname'),
            ('nationality', 'nationality'),
            ('citizen_id', 'citizenId'),
            ('cccd_date', 'cccdDate'),
            ('cccd_place', 'cccdPlace'),
            ('personal_id', 'personalId'),
            ('passport', 'passport'),
            ('passport_date', 'passportDate'),
            ('passport_place', 'passportPlace'),
            ('occupation', 'occupation'),
            ('organization', 'organization'),
            # Address info
            ('permanent_province', 'permanentProvince'),
            ('permanent_ward', 'permanentWard'),
            ('permanent_hamlet', 'permanentHamlet'),
            ('permanent_street', 'permanentStreet'),
            ('hometown_province', 'hometownProvince'),
            ('hometown_ward', 'hometownWard'),
            ('hometown_hamlet', 'hometownHamlet'),
            ('current_ward', 'currentWard'),
            ('current_hamlet', 'currentHamlet'),
            ('birthplace_province', 'birthplaceProvince'),
            ('birthplace_ward', 'birthplaceWard'),
            ('birth_cert_province', 'birthCertProvince'),
            ('birth_cert_ward', 'birthCertWard'),
            # Health info
            ('height', 'height'),
            ('weight', 'weight'),
            ('eye_diseases', 'eyeDiseases'),  # Fixed: eyeDiseases (plural) from frontend
            ('swimming_skill', 'swimmingSkill'),
            # Device info
            ('smartphone', 'smartphone'),
            ('computer', 'computer'),
            # Family info
            ('father_ethnicity', 'fatherEthnicity'),
            ('father_birth_year', 'fatherBirthYear'),
            ('father_phone', 'fatherPhone'),
            ('father_cccd', 'fatherCCCD'),
            ('mother_ethnicity', 'motherEthnicity'),
            ('mother_birth_year', 'motherBirthYear'),
            ('mother_phone', 'motherPhone'),
            ('mother_cccd', 'motherCCCD'),
            ('guardian_name', 'guardianName'),
            ('guardian_job', 'guardianJob'),
            ('guardian_birth_year', 'guardianBirthYear'),
            ('guardian_phone', 'guardianPhone'),
            ('guardian_cccd', 'guardianCCCD'),
            ('guardian_gender', 'guardianGender')
        ]

        def normalize_value(db_col, val):
            if db_col == 'eye_diseases':
                if isinstance(val, list):
                    return ','.join(val)
                return val
            elif db_col in ['ngay_sinh', 'cccd_date', 'passport_date']:
                # Convert dd/mm/yyyy to yyyy-mm-dd for PostgreSQL
                if val and isinstance(val, str) and val.strip():
                    try:
                        # Handle dd/mm/yyyy format
                        if '/' in val:
                            parts = val.split('/')
                            if len(parts) == 3:
                                day, month, year = parts
                                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        # Handle yyyy-mm-dd format (already correct)
                        elif '-' in val and len(val) == 10:
                            return val
                    except:
                        pass
                # Return None for empty or invalid dates
                return None
            elif db_col in ['height', 'weight', 'father_birth_year', 'mother_birth_year', 'guardian_birth_year']:
                # Handle integer fields - convert empty string to None
                if val and isinstance(val, str) and val.strip():
                    try:
                        return int(val)
                    except ValueError:
                        return None
                elif isinstance(val, int):
                    return val
                return None
            return val

        payload = {}
        for db_col, json_key in col_map:
            payload[db_col] = normalize_value(db_col, data.get(json_key))
        
        # Extract grade (khoi) from class if not provided directly
        if not payload.get('khoi') and payload.get('lop'):
            class_value = payload['lop']
            if class_value and len(class_value) >= 2:
                khoi = class_value[:2]  # Extract first 2 characters (10, 11, 12)
                payload['khoi'] = khoi

        if existing:
            update_cols = [c for c, _ in col_map if c != 'email']
            if DB_CONFIG['type'] == 'postgresql':
                set_clause = ', '.join([f"{c} = %s" for c in update_cols])
                set_clause = f"{set_clause}, created_at = CURRENT_TIMESTAMP"
                values = [payload[c] for c in update_cols]
                values.append(payload['email'])
                cursor.execute(f"UPDATE students SET {set_clause} WHERE email = %s", values)
            else:
                set_clause = ', '.join([f"{c} = ?" for c in update_cols])
                set_clause = f"{set_clause}, created_at = CURRENT_TIMESTAMP"
                values = [payload[c] for c in update_cols]
                values.append(payload['email'])
                cursor.execute(f"UPDATE students SET {set_clause} WHERE email = ?", values)
        else:
            insert_cols = [c for c, _ in col_map]
            if DB_CONFIG['type'] == 'postgresql':
                placeholders = ', '.join(['%s'] * len(insert_cols))
                values = [payload[c] for c in insert_cols]
                cursor.execute(
                    f"INSERT INTO students ({', '.join(insert_cols)}) VALUES ({placeholders})",
                    values
                )
            else:
                placeholders = ', '.join(['?'] * len(insert_cols))
                values = [payload[c] for c in insert_cols]
                cursor.execute(
                    f"INSERT INTO students ({', '.join(insert_cols)}) VALUES ({placeholders})",
                    values
                )

        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': 'Dữ liệu đã được lưu thành công!'})

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'}), 500

@app.errorhandler(405)
def method_not_allowed(e):
    path = request.path or ''
    if path.startswith('/api/'):
        return jsonify({'success': False, 'message': 'Phương thức không được phép cho endpoint này'}), 405
    return e

@app.route('/api/students', methods=['GET'])
def get_students():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        search = request.args.get('search', '').strip()

        if page < 1:
            page = 1
        if limit < 1 or limit > 100:
            limit = 50

        offset = (page - 1) * limit

        conn = get_db_connection()
        # Only set row_factory for SQLite
        if DB_CONFIG['type'] == 'sqlite':
            conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Check available columns for debugging
        if DB_CONFIG['type'] == 'postgresql':
            cursor.execute("""
                SELECT column_name FROM information_schema.columns 
                WHERE table_name = 'students'
            """)
            available_columns = [row[0] for row in cursor.fetchall()]
            print(f"[ADMIN API] PostgreSQL columns: {available_columns}")
            has_eye_diseases = 'eye_diseases' in available_columns
            print(f"[ADMIN API] Has eye_diseases column: {has_eye_diseases}")
        else:
            cursor.execute('PRAGMA table_info(students)')
            available_columns = [col[1] for col in cursor.fetchall()]
            print(f"[ADMIN API] SQLite columns: {available_columns}")
            has_eye_diseases = 'eye_diseases' in available_columns

        if DB_CONFIG['type'] == 'postgresql':
            # PostgreSQL syntax with placeholders - only include eye_diseases if it exists
            if has_eye_diseases:
                base_query = """
                SELECT id, email, ho_ten as full_name, email as nickname, lop as class, ngay_sinh as birth_date, gioi_tinh as gender,
                       sdt as phone, created_at, eye_diseases, tinh_thanh as current_province
                FROM students
                """
            else:
                print("[ADMIN API] eye_diseases column not found in PostgreSQL, excluding from query")
                base_query = """
                SELECT id, email, ho_ten as full_name, email as nickname, lop as class, ngay_sinh as birth_date, gioi_tinh as gender,
                       sdt as phone, created_at, tinh_thanh as current_province
                FROM students
                """
            count_query = "SELECT COUNT(*) as total FROM students"

            where_clause = ""
            params = []
            if search:
                where_clause = """
                WHERE ho_ten ILIKE %s OR email ILIKE %s OR lop ILIKE %s
                OR sdt ILIKE %s
                """
                search_param = f"%{search}%"
                params = [search_param, search_param, search_param, search_param]

            # Get total count
            cursor.execute(count_query + where_clause, params)
            total = cursor.fetchone()[0]

            # Get records with pagination
            query = base_query + where_clause + " ORDER BY created_at DESC LIMIT %s OFFSET %s"
            cursor.execute(query, params + [limit, offset])
            
        else:
            # SQLite syntax
            base_query = """
            SELECT id, email, full_name, nickname, class, birth_date, gender,
                   phone, created_at, eye_diseases, current_province
            FROM students
            """
            count_query = "SELECT COUNT(*) as total FROM students"

            where_clause = ""
            params = []
            if search:
                where_clause = """
                WHERE full_name LIKE ? OR email LIKE ? OR class LIKE ?
                OR phone LIKE ? OR nickname LIKE ?
                """
                search_param = f"%{search}%"
                params = [search_param] * 5

            # Get total count
            cursor.execute(count_query + where_clause, params)
            total = cursor.fetchone()[0]

            # Get records with pagination
            query = base_query + where_clause + " ORDER BY created_at DESC LIMIT ? OFFSET ?"
            cursor.execute(query, params + [limit, offset])

        # Process results
        students = []
        rows = cursor.fetchall()
        
        if DB_CONFIG['type'] == 'postgresql':
            # Convert PostgreSQL results to dict
            column_names = [desc[0] for desc in cursor.description]
            for row in rows:
                student = dict(zip(column_names, row))
                # Add field mappings for frontend compatibility
                student['eyeDiseases'] = student.get('eye_diseases', '')
                student['tinh_thanh'] = student.get('current_province', '') or student.get('tinh_thanh', '')
                # Also ensure original field names exist for fallback
                if 'eye_diseases' not in student:
                    student['eye_diseases'] = student.get('eyeDiseases', '')
                students.append(student)
        else:
            # SQLite with row_factory
            for row in rows:
                student = dict(row)
                # Add field mappings for frontend compatibility
                student['eyeDiseases'] = student.get('eye_diseases', '')
                student['tinh_thanh'] = student.get('current_province', '')
                # Also ensure original field names exist for fallback
                if 'eye_diseases' not in student:
                    student['eye_diseases'] = student.get('eyeDiseases', '')
                students.append(student)

        conn.close()

        total_pages = math.ceil(total / limit)

        return jsonify({
            'data': students,
            'students': students,
            'pagination': {
                'current_page': page,
                'total_pages': total_pages,
                'total_records': total,
                'per_page': limit,
                'limit': limit,
                'has_next': page < total_pages,
                'has_prev': page > 1
            },
            'search': search
        })

    except Exception as e:
        print(f"[API ERROR] get_students: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/<int:student_id>', methods=['GET'])
def get_student_detail(student_id):
    try:
        conn = get_db_connection()
        # Only set row_factory for SQLite
        if DB_CONFIG['type'] == 'sqlite':
            conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        if DB_CONFIG['type'] == 'postgresql':
            cursor.execute('SELECT * FROM students WHERE id = %s', (student_id,))
        else:
            cursor.execute('SELECT * FROM students WHERE id = ?', (student_id,))
            
        row = cursor.fetchone()
        conn.close()

        if not row:
            return jsonify({'error': 'Không tìm thấy học sinh'}), 404

        if DB_CONFIG['type'] == 'postgresql':
            # Convert PostgreSQL result to dict
            column_names = [desc[0] for desc in cursor.description]
            student = dict(zip(column_names, row))
            print(f"[STUDENT DETAIL] PostgreSQL columns: {column_names}")
            print(f"[STUDENT DETAIL] Has eye_diseases: {'eye_diseases' in column_names}")
            if 'eye_diseases' in student:
                print(f"[STUDENT DETAIL] eye_diseases value: '{student['eye_diseases']}'")
        else:
            # SQLite with row_factory
            student = dict(row)
            print(f"[STUDENT DETAIL] SQLite columns: {list(student.keys())}")
            if 'eye_diseases' in student:
                print(f"[STUDENT DETAIL] eye_diseases value: '{student['eye_diseases']}'")

        # Debug field mapping
        eye_diseases_value = student.get('eye_diseases', '')
        print(f"[STUDENT DETAIL] Mapping eye_diseases '{eye_diseases_value}' to eyeDiseases")

        if student.get('permanent_street') and student.get('permanent_hamlet') and student.get('permanent_ward') and student.get('permanent_province'):
            student['permanent_address'] = f"{student['permanent_street']}, {student['permanent_hamlet']}, {student['permanent_ward']}, {student['permanent_province']}"
        else:
            student['permanent_address'] = None

        if student.get('current_address_detail') and student.get('current_hamlet') and student.get('current_ward') and student.get('current_province'):
            student['temporary_address'] = f"{student['current_address_detail']}, {student['current_hamlet']}, {student['current_ward']}, {student['current_province']}"
        else:
            student['temporary_address'] = None

        student['id_number'] = student.get('citizen_id') or student.get('personal_id')
        
        # Add field mappings for frontend compatibility
        student['eyeDiseases'] = student.get('eye_diseases', '')
        student['tinh_thanh'] = student.get('current_province', '') or student.get('tinh_thanh', '')
        
        # CRITICAL: Also keep the original database field names for fallback
        if 'eye_diseases' not in student:
            student['eye_diseases'] = student.get('eyeDiseases', '')
        if 'current_province' not in student and 'tinh_thanh' in student:
            student['current_province'] = student.get('tinh_thanh', '')
        
        print(f"[STUDENT DETAIL] Final eyeDiseases value: '{student['eyeDiseases']}'")
        print(f"[STUDENT DETAIL] Final eye_diseases value: '{student.get('eye_diseases', '')}'")
        print(f"[STUDENT DETAIL] Final tinh_thanh value: '{student['tinh_thanh']}'")

        return jsonify(student)

    except Exception as e:
        print(f"[API ERROR] get_student_detail: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/debug/schema', methods=['GET'])
def debug_schema():
    """Debug endpoint to check database schema"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if DB_CONFIG['type'] == 'postgresql':
            cursor.execute("""
                SELECT column_name, data_type, is_nullable 
                FROM information_schema.columns 
                WHERE table_name = 'students'
                ORDER BY ordinal_position
            """)
            columns = cursor.fetchall()
            schema_info = {
                'database_type': 'PostgreSQL',
                'columns': [{'name': col[0], 'type': col[1], 'nullable': col[2]} for col in columns],
                'total_columns': len(columns)
            }
        else:
            cursor.execute('PRAGMA table_info(students)')
            columns = cursor.fetchall()
            schema_info = {
                'database_type': 'SQLite',
                'columns': [{'name': col[1], 'type': col[2], 'nullable': not col[3]} for col in columns],
                'total_columns': len(columns)
            }
        
        # Check for eye_diseases specifically
        has_eye_diseases = any(col['name'] == 'eye_diseases' for col in schema_info['columns'])
        schema_info['has_eye_diseases'] = has_eye_diseases
        
        # Get sample data if eye_diseases exists
        if has_eye_diseases:
            cursor.execute('SELECT id, full_name, eye_diseases FROM students WHERE eye_diseases IS NOT NULL AND eye_diseases != \'\' LIMIT 5')
            sample_data = cursor.fetchall()
            schema_info['sample_eye_diseases'] = [{'id': row[0], 'name': row[1], 'eye_diseases': row[2]} for row in sample_data]
        
        conn.close()
        return jsonify(schema_info)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student-by-email', methods=['GET'])
def find_student_by_email():
    try:
        email = request.args.get('email')
        if not email:
            return jsonify({'error': 'Thiếu tham số email'}), 400

        conn = get_db_connection()
        # Only set row_factory for SQLite
        if DB_CONFIG['type'] == 'sqlite':
            conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        if DB_CONFIG['type'] == 'postgresql':
            cursor.execute('''
                SELECT * FROM students WHERE email = %s
                ORDER BY created_at DESC, id DESC LIMIT 1
            ''', (email,))
        else:
            cursor.execute('''
                SELECT * FROM students WHERE email = ?
                ORDER BY datetime(created_at) DESC, id DESC LIMIT 1
            ''', (email,))
            
        row = cursor.fetchone()
        conn.close()

        if not row:
            return jsonify({'student': None}), 200

        if DB_CONFIG['type'] == 'postgresql':
            # Convert PostgreSQL result to dict
            column_names = [desc[0] for desc in cursor.description]
            student = dict(zip(column_names, row))
        else:
            # SQLite with row_factory
            student = {k: row[k] for k in row.keys()}
            
        return jsonify({'student': student})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    try:
        # Lấy parameters từ request
        grade = request.args.get('grade')  # '10', '11', '12'
        classes = request.args.get('classes')  # '10A1,10A2,11B3'
        province = request.args.get('province')  # Province filter
        ethnicity = request.args.get('ethnicity')  # Ethnicity filter
        font_size = int(request.args.get('fontSize', '11'))  # Font size parameter
        
        print(f"[EXCEL] Starting export - Grade: {grade}, Classes: {classes}, Province: {province}, Ethnicity: {ethnicity}, FontSize: {font_size}")

        conn = get_db_connection()
        conn.execute('PRAGMA temp_store = MEMORY')
        conn.execute('PRAGMA cache_size = 10000')

        # Xây dựng câu query với filter
        base_query = 'SELECT * FROM students'
        where_conditions = []
        query_params = []

        if grade:
            # Filter theo khối học - sử dụng SUBSTR để chính xác
            if DB_CONFIG['type'] == 'postgres':
                where_conditions.append("substring(class, 1, LENGTH(?)) = ?")
            else:
                where_conditions.append("SUBSTR(class, 1, LENGTH(?)) = ?")
            query_params.extend([grade, grade])
            print(f"[EXCEL] Filtering by grade: {grade}")
        elif classes:
            # Filter theo lớp cụ thể
            class_list = [cls.strip() for cls in classes.split(',')]
            placeholders = ','.join(['?' for _ in class_list])
            where_conditions.append(f"class IN ({placeholders})")
            query_params.extend(class_list)
            print(f"[EXCEL] Filtering by classes: {class_list}")

        # Filter theo tỉnh thường trú
        if province:
            where_conditions.append("permanent_province = ?")
            query_params.append(province)
            print(f"[EXCEL] Filtering by province: {province}")

        # Filter theo dân tộc
        if ethnicity:
            where_conditions.append("ethnicity = ?")
            query_params.append(ethnicity)
            print(f"[EXCEL] Filtering by ethnicity: {ethnicity}")

        # Tạo câu query hoàn chỉnh
        if where_conditions:
            query = f"{base_query} WHERE {' AND '.join(where_conditions)} ORDER BY id ASC"
        else:
            query = f"{base_query} ORDER BY id ASC"
            
        print(f"[EXCEL] Query: {query}")
        print(f"[EXCEL] Params: {query_params}")

        # Đọc dữ liệu với chunks để xử lý dataset lớn
        if query_params:
            # Pandas không hỗ trợ tham số với chunksize, nên ta phải đọc trực tiếp
            cursor = conn.cursor()
            cursor.execute(query, query_params)
            rows = cursor.fetchall()
            
            # Lấy column names
            column_names = [description[0] for description in cursor.description]
            
            # Tạo DataFrame
            df_final = pd.DataFrame(rows, columns=column_names)
            total_records = len(df_final)
            print(f"[EXCEL] Filtered records: {total_records}")
        else:
            # Không có filter, đọc tất cả với chunks
            df = pd.read_sql_query(query, conn, chunksize=1000)
            df_list = []
            total_records = 0
            for chunk in df:
                df_list.append(chunk)
                total_records += len(chunk)
                print(f"[EXCEL] Processed {total_records} records...")
            
            if not df_list:
                conn.close()
                return jsonify({'error': 'Không có dữ liệu để xuất'}), 400
                
            df_final = pd.concat(df_list, ignore_index=True)

        conn.close()

        if df_final.empty:
            return jsonify({'error': 'Không có dữ liệu phù hợp để xuất'}), 400

        print(f"[EXCEL] Total records to export: {total_records}")

        # Tạo filename phù hợp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if grade:
            filename = f'danh_sach_hoc_sinh_khoi_{grade}_{timestamp}.xlsx'
        elif classes:
            class_list = [cls.strip() for cls in classes.split(',')]
            if len(class_list) == 1:
                filename = f'danh_sach_hoc_sinh_lop_{class_list[0]}_{timestamp}.xlsx'
            else:
                filename = f'danh_sach_hoc_sinh_{len(class_list)}_lop_{timestamp}.xlsx'
        else:
            filename = f'danh_sach_hoc_sinh_tat_ca_{timestamp}.xlsx'

        column_mapping = {
            'id': 'STT',
            'email': 'Email',
            'full_name': 'Họ và tên',
            'nickname': 'Tên gọi khác',
            'class': 'Lớp',
            'birth_date': 'Ngày sinh',
            'gender': 'Giới tính',
            'ethnicity': 'Dân tộc',
            'dan_toc': 'Dân tộc',  # Mapping cho PostgreSQL
            'nationality': 'Quốc tịch',
            'religion': 'Tôn giáo',
            'phone': 'Số điện thoại',
            'citizen_id': 'Số CCCD',
            'cccd_date': 'Ngày cấp CCCD',
            'cccd_place': 'Nơi cấp CCCD',
            'personal_id': 'Mã định danh',
            'passport': 'Số hộ chiếu',
            'passport_date': 'Ngày cấp hộ chiếu',
            'passport_place': 'Nơi cấp hộ chiếu',
            'organization': 'Đoàn/Đội',
            'permanent_province': 'Tỉnh thường trú',
            'permanent_ward': 'Phường thường trú',
            'permanent_hamlet': 'Khu phố thường trú',
            'permanent_street': 'Địa chỉ thường trú',
            'hometown_province': 'Tỉnh quê quán',
            'hometown_ward': 'Phường quê quán',
            'hometown_hamlet': 'Khu phố quê quán',
            'birth_cert_province': 'Tỉnh cấp giấy khai sinh',
            'birth_cert_ward': 'Phường cấp giấy khai sinh',
            'birthplace_province': 'Tỉnh nơi sinh',
            'birthplace_ward': 'Phường nơi sinh',
            'current_address_detail': 'Địa chỉ hiện tại',
            'current_province': 'Tỉnh hiện tại',
            'current_ward': 'Phường hiện tại',
            'current_hamlet': 'Khu phố hiện tại',
            'height': 'Chiều cao (cm)',
            'weight': 'Cân nặng (kg)',
            'eye_diseases': 'Tật khúc xạ (mắt)',
            'swimming_skill': 'Kỹ năng bơi',
            'smartphone': 'Điện thoại thông minh',
            'computer': 'Máy tính',
            'father_ethnicity': 'Dân tộc của cha',
            'mother_ethnicity': 'Dân tộc của mẹ',
            'father_name': 'Họ tên cha',
            'father_job': 'Nghề nghiệp cha',
            'father_birth_year': 'Năm sinh cha',
            'father_phone': 'SĐT cha',
            'father_cccd': 'CCCD cha',
            'mother_name': 'Họ tên mẹ',
            'mother_job': 'Nghề nghiệp mẹ',
            'mother_birth_year': 'Năm sinh mẹ',
            'mother_phone': 'SĐT mẹ',
            'mother_cccd': 'CCCD mẹ',
            'guardian_name': 'Họ tên người giám hộ',
            'guardian_job': 'Nghề nghiệp người giám hộ',
            'guardian_birth_year': 'Năm sinh người giám hộ',
            'guardian_phone': 'SĐT người giám hộ',
            'guardian_cccd': 'CCCD người giám hộ',
            'guardian_gender': 'Giới tính người giám hộ',
            'created_at': 'Thời gian nộp kê khai'
        }

        df_export = df_final.rename(columns=column_mapping)

        order_keys = [
            'id',
            'email','full_name','nickname','class','birth_date','gender','ethnicity','nationality','religion','phone',
            'citizen_id','cccd_date','cccd_place','personal_id','passport','passport_date','passport_place','organization',
            'permanent_province','permanent_ward','permanent_hamlet','permanent_street',
            'hometown_province','hometown_ward','hometown_hamlet',
            'birth_cert_province','birth_cert_ward','birthplace_province','birthplace_ward',
            'current_address_detail','current_province','current_ward','current_hamlet',
            'height','weight','eye_diseases','swimming_skill',
            'smartphone','computer',
            'father_name','father_ethnicity','father_job','father_birth_year','father_phone','father_cccd',
            'mother_name','mother_ethnicity','mother_job','mother_birth_year','mother_phone','mother_cccd',
            'guardian_name','guardian_job','guardian_birth_year','guardian_phone','guardian_cccd','guardian_gender',
            'created_at'
        ]
        order_vn = [column_mapping.get(k, k) for k in order_keys]
        ordered_present = [c for c in order_vn if c in df_export.columns]
        others = [c for c in df_export.columns if c not in ordered_present]
        df_export = df_export[ordered_present + others]

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active
            ws.title = "Danh sách học sinh"

            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws.append(r)

            header_font = Font(bold=True, color="FFFFFF", size=font_size + 1)  # Dynamic header font
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            content_font = Font(color="000000", size=font_size)  # Dynamic content font
            content_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            content_alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = content_font
                    cell.fill = content_fill
                    cell.alignment = content_alignment
                    cell.border = thin_border

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        # Handle Vietnamese characters properly
                        cell_length = len(cell_value)
                        # Add extra padding for Vietnamese characters
                        if any(ord(char) > 127 for char in cell_value):
                            cell_length = int(cell_length * 1.2)  # 20% extra for Vietnamese
                        
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

                # Set minimum width of 12, maximum of 80, with 3 characters padding
                adjusted_width = min(max(max_length + 3, 12), 80)
                ws.column_dimensions[column_letter].width = adjusted_width

            ws.row_dimensions[1].height = 25

            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 20

            wb.save(filename)

        except ImportError:
            return jsonify({'error': "Thiếu thư viện 'openpyxl'. Vui lòng chạy start.bat hoặc cài đặt bằng lệnh: .\\.venv\\Scripts\\pip.exe install openpyxl"}), 500
        except Exception as e:
            print(f"[EXCEL] Warning: Styling failed, using basic export: {str(e)}")
            df_export.to_excel(filename, index=False)

        return send_file_with_cleanup(filename, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Enhanced export endpoints for different formats
@app.route('/api/export-xlsx', methods=['GET'])
def export_xlsx():
    """Enhanced XLSX export with more options"""
    try:
        # Get all parameters from request
        export_type = request.args.get('type', 'all')  # all, grade, class, custom
        grade = request.args.get('grade')  
        classes = request.args.get('classes')  
        province = request.args.get('province')  # Province filter
        ethnicity = request.args.get('ethnicity')  # Ethnicity filter
        ethnicity = request.args.get('ethnicity')  # Ethnicity filter
        title = request.args.get('title', 'Danh sách học sinh THPT Dĩ An')
        include_stats = request.args.get('includeStats') == 'true'
        include_timestamp = request.args.get('includeTimestamp') == 'true'
        sort_by_class = request.args.get('sortByClass') == 'true'
        sort_by_name = request.args.get('sortByName') == 'true'
        hide_empty_fields = request.args.get('hideEmptyFields') == 'true'
        theme_color = request.args.get('themeColor', 'blue')
        font_size = int(request.args.get('fontSize', '11'))  # Get font size parameter
        
        # Auto detect export type if not provided
        if export_type == 'all':
            if grade:
                export_type = 'grade'
            elif classes:
                export_type = 'class'
            elif (request.args.get('gender') or 
                  request.args.get('fromYear') or 
                  request.args.get('toYear') or 
                  request.args.get('hasPhone') or
                  province or ethnicity):
                export_type = 'custom'
        
        print(f"[XLSX] Export type: {export_type}, Grade: {grade}, Classes: {classes}, Province: {province}, Ethnicity: {ethnicity}")
        
        # Debug custom filters
        if export_type == 'custom':
            gender = request.args.get('gender')
            from_year = request.args.get('fromYear')
            to_year = request.args.get('toYear') 
            has_phone = request.args.get('hasPhone')
            print(f"[XLSX] Custom filters detected - Gender: {gender}, Years: {from_year}-{to_year}, HasPhone: {has_phone}")

        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Auto-detect database schema
        if DB_CONFIG['type'] == 'postgresql':
            cursor.execute("""
                SELECT column_name FROM information_schema.columns 
                WHERE table_name = 'students'
            """)
            existing_columns = [row[0] for row in cursor.fetchall()]
        else:
            cursor.execute('PRAGMA table_info(students)')
            existing_columns = [col[1] for col in cursor.fetchall()]
        
        print(f"[EXPORT] Found {len(existing_columns)} columns in database")
        print(f"[EXPORT] Available columns: {existing_columns}")
        
        # Check specifically for eye_diseases column
        has_eye_diseases = 'eye_diseases' in existing_columns
        print(f"[EXPORT] Has eye_diseases column: {has_eye_diseases}")
        
        # Map old column names to new ones
        column_mapping = {
            'ho_ten': 'full_name',
            'lop': 'class', 
            'ngay_sinh': 'birth_date',
            'gioi_tinh': 'gender',
            'dan_toc': 'ethnicity',
            'sdt': 'phone',
            'dia_chi': 'current_address_detail',
            'tinh_thanh': 'current_province',
            'ho_ten_cha': 'father_name',
            'nghe_nghiep_cha': 'father_job',
            'ho_ten_me': 'mother_name',
            'nghe_nghiep_me': 'mother_job'
        }
        
        # Determine which columns to use based on what exists
        use_old_schema = 'ho_ten' in existing_columns
        
        # Always use new schema for export to avoid duplicate columns
        print("[EXPORT] Using new schema for clean export (full_name, class, ethnicity)")
        basic_columns = [
            'id', 'email', 'full_name', 'nickname', 'class', 'birth_date', 'gender', 'ethnicity', 
            'nationality', 'religion', 'phone', 'citizen_id', 'cccd_date', 'cccd_place', 
            'personal_id', 'passport', 'passport_date', 'passport_place', 'organization',
            'permanent_province', 'permanent_ward', 'permanent_hamlet', 'permanent_street',
            'hometown_province', 'hometown_ward', 'hometown_hamlet', 
            'current_address_detail', 'current_province', 'current_ward', 'current_hamlet',
            'birthplace_province', 'birthplace_ward', 'birth_cert_province', 'birth_cert_ward',
            'height', 'weight', 'eye_diseases', 'swimming_skill', 'smartphone', 'computer',
            'father_name', 'father_ethnicity', 'father_job', 'father_birth_year', 'father_phone', 'father_cccd',
            'mother_name', 'mother_ethnicity', 'mother_job', 'mother_birth_year', 'mother_phone', 'mother_cccd',
            'guardian_name', 'guardian_job', 'guardian_birth_year', 'guardian_phone', 'guardian_cccd', 'guardian_gender',
            'created_at'
        ]
        
        # Determine column names for queries based on actual schema
        if use_old_schema:
            class_column = 'lop'
            name_column = 'ho_ten'
            gender_column = 'gioi_tinh'
            ethnicity_column = 'dan_toc'
            phone_column = 'sdt'
        else:
            class_column = 'class'
            name_column = 'full_name'
            gender_column = 'gender'
            ethnicity_column = 'ethnicity'
            phone_column = 'phone'
            
        # Always include essential columns regardless of schema
        essential_old_columns = [
            'id', 'ho_ten', 'ngay_sinh', 'gioi_tinh', 'dan_toc', 'lop', 'khoi', 'sdt', 
            'ton_giao', 'dia_chi', 'tinh_thanh', 'ho_ten_cha', 'nghe_nghiep_cha', 
            'ho_ten_me', 'nghe_nghiep_me'
        ]
        essential_new_columns = [
            'id', 'email', 'full_name', 'birth_date', 'gender', 'ethnicity', 'class', 'phone',
            'religion', 'current_address_detail', 'current_province', 'father_name', 'father_job',
            'mother_name', 'mother_job', 'height', 'weight', 'eye_diseases', 'swimming_skill',
            'smartphone', 'computer', 'nationality'
        ]
        
        # Combine both old and new essential columns with all other columns
        all_possible_columns = list(set(basic_columns + essential_old_columns + essential_new_columns))
        
        # Filter columns to only include existing ones
        available_columns = [col for col in all_possible_columns if col in existing_columns]
        missing_columns = [col for col in all_possible_columns if col not in existing_columns]
            
        print(f"[EXPORT] Using {len(available_columns)} available columns")
        print(f"[EXPORT] Missing columns: {missing_columns}")
        
        column_list = ', '.join(available_columns)
        base_query = f'SELECT {column_list} FROM students'
        where_conditions = []
        query_params = []

        if export_type == 'grade' and grade:
            # Filter theo khối học chính xác - chỉ lấy các lớp thuộc khối đó
            placeholder = get_placeholder()
            if DB_CONFIG['type'] == 'postgres':
                where_conditions.append(f"substring({class_column}, 1, LENGTH({placeholder})) = {placeholder}")
            else:
                where_conditions.append(f"SUBSTR({class_column}, 1, LENGTH({placeholder})) = {placeholder}")
            query_params.extend([grade, grade])
            print(f"[XLSX] Filtering by grade: {grade}")
        elif export_type == 'class' and classes:
            class_list = [cls.strip() for cls in classes.split(',')]
            placeholder = get_placeholder()
            placeholders = ','.join([placeholder for _ in class_list])
            where_conditions.append(f"{class_column} IN ({placeholders})")
            query_params.extend(class_list)
            print(f"[XLSX] Filtering by classes: {class_list}")
        elif export_type == 'custom':
            # Handle custom filters
            gender = request.args.get('gender')
            has_phone = request.args.get('hasPhone') == 'true'
            
            print(f"[XLSX] Custom filters - Gender: {gender}, HasPhone: {has_phone}")
            
            if gender:
                gender_list = [g.strip() for g in gender.split(',')]
                placeholder = get_placeholder()
                gender_placeholders = ','.join([placeholder for _ in gender_list])
                where_conditions.append(f"{gender_column} IN ({gender_placeholders})")
                query_params.extend(gender_list)
                
            if has_phone:
                where_conditions.append(f"{phone_column} IS NOT NULL AND {phone_column} != ''")
        
        # Apply province and ethnicity filters for ALL export types
        if province:
            placeholder = get_placeholder()
            # Flexible province matching - case insensitive and partial match
            where_conditions.append(f"LOWER(permanent_province) LIKE LOWER({placeholder})")
            query_params.append(f"%{province}%")
            print(f"[XLSX] Filtering by province: %{province}%")

        if ethnicity:
            placeholder = get_placeholder()
            # Flexible ethnicity matching - case insensitive and partial match
            where_conditions.append(f"LOWER({ethnicity_column}) LIKE LOWER({placeholder})")
            query_params.append(f"%{ethnicity}%")
            print(f"[XLSX] Filtering by ethnicity: %{ethnicity}%")

        # Build final query
        if where_conditions:
            query = f"{base_query} WHERE {' AND '.join(where_conditions)}"
        else:
            query = base_query
            
        # Add sorting
        if sort_by_class:
            query += f" ORDER BY {class_column}, {name_column}"
        elif sort_by_name:
            query += f" ORDER BY {name_column}, {class_column}"
        else:
            query += " ORDER BY id ASC"

        # Execute query
        if query_params:
            cursor = conn.cursor()
            cursor.execute(query, query_params)
            rows = cursor.fetchall()
            column_names = [description[0] for description in cursor.description]
            df_final = pd.DataFrame(rows, columns=column_names)
        else:
            df_final = pd.read_sql_query(query, conn)

        conn.close()

        if df_final.empty:
            return jsonify({'error': 'Không có dữ liệu phù hợp để xuất'}), 400

        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if export_type == 'grade' and grade:
            filename = f'danh_sach_hoc_sinh_khoi_{grade}_{timestamp}.xlsx'
        elif export_type == 'class' and classes:
            class_list = [cls.strip() for cls in classes.split(',')]
            if len(class_list) == 1:
                filename = f'danh_sach_hoc_sinh_lop_{class_list[0]}_{timestamp}.xlsx'
            else:
                filename = f'danh_sach_hoc_sinh_{len(class_list)}_lop_{timestamp}.xlsx'
        elif export_type == 'custom':
            filename = f'danh_sach_hoc_sinh_tuy_chinh_{timestamp}.xlsx'
        else:
            filename = f'danh_sach_hoc_sinh_tat_ca_{timestamp}.xlsx'

        # Column mapping - using actual database column names with old->new schema mapping
        column_mapping = {
            'id': 'STT',
            'email': 'Email', 
            'full_name': 'Họ và tên',
            'nickname': 'Tên gọi khác',
            'class': 'Lớp',
            'birth_date': 'Ngày sinh',
            'gender': 'Giới tính',
            'ethnicity': 'Dân tộc',
            'nationality': 'Quốc tịch',
            'religion': 'Tôn giáo',
            'phone': 'Số điện thoại',
            'citizen_id': 'Số CCCD',
            'cccd_date': 'Ngày cấp CCCD',
            'cccd_place': 'Nơi cấp CCCD',
            'personal_id': 'Mã định danh',
            'passport': 'Số hộ chiếu',
            'passport_date': 'Ngày cấp hộ chiếu',
            'passport_place': 'Nơi cấp hộ chiếu',
            'organization': 'Đoàn/Đội',
            'permanent_province': 'Tỉnh thường trú',
            'permanent_ward': 'Phường thường trú',
            'permanent_hamlet': 'Khu phố thường trú',
            'permanent_street': 'Địa chỉ thường trú',
            'hometown_province': 'Tỉnh quê quán',
            'hometown_ward': 'Phường quê quán',
            'hometown_hamlet': 'Khu phố quê quán',
            'current_address_detail': 'Địa chỉ chi tiết hiện tại',
            'current_province': 'Tỉnh hiện tại',
            'current_ward': 'Phường hiện tại',
            'current_hamlet': 'Khu phố hiện tại',
            'birthplace_province': 'Tỉnh nơi sinh',
            'birthplace_ward': 'Phường nơi sinh',
            'birth_cert_province': 'Tỉnh cấp giấy khai sinh',
            'birth_cert_ward': 'Phường cấp giấy khai sinh',
            'height': 'Chiều cao (cm)',
            'weight': 'Cân nặng (kg)',
            'eye_diseases': 'Tật khúc xạ (mắt)',
            'swimming_skill': 'Kỹ năng bơi',
            'smartphone': 'Điện thoại thông minh',
            'computer': 'Máy tính',
            'father_name': 'Họ tên cha',
            'father_ethnicity': 'Dân tộc của cha',
            'father_job': 'Nghề nghiệp cha',
            'father_birth_year': 'Năm sinh cha',
            'father_phone': 'SĐT cha',
            'father_cccd': 'CCCD cha',
            'mother_name': 'Họ tên mẹ',
            'mother_ethnicity': 'Dân tộc của mẹ',
            'mother_job': 'Nghề nghiệp mẹ',
            'mother_birth_year': 'Năm sinh mẹ',
            'mother_phone': 'SĐT mẹ',
            'mother_cccd': 'CCCD mẹ',
            'guardian_name': 'Họ tên người giám hộ',
            'guardian_job': 'Nghề nghiệp người giám hộ',
            'guardian_birth_year': 'Năm sinh người giám hộ',
            'guardian_phone': 'SĐT người giám hộ',
            'guardian_cccd': 'CCCD người giám hộ',
            'guardian_gender': 'Giới tính người giám hộ',
            'created_at': 'Thời gian nộp kê khai',
            # Old schema mapping - map old columns to new Vietnamese names
            'ho_ten': 'Họ và tên',
            'ngay_sinh': 'Ngày sinh', 
            'gioi_tinh': 'Giới tính',
            'dan_toc': 'Dân tộc',
            'lop': 'Lớp',
            'khoi': 'Khối',
            'sdt': 'Số điện thoại',
            'ton_giao': 'Tôn giáo',
            'dia_chi': 'Địa chỉ chi tiết hiện tại',
            'tinh_thanh': 'Tỉnh hiện tại',
            'ho_ten_cha': 'Họ tên cha',
            'nghe_nghiep_cha': 'Nghề nghiệp cha',
            'ho_ten_me': 'Họ tên mẹ',
            'nghe_nghiep_me': 'Nghề nghiệp mẹ'
        }

        df_export = df_final.rename(columns=column_mapping)
        
        # Handle duplicate columns if both old and new schema exist
        # Ensure we have the essential columns with correct data
        if use_old_schema:
            print("[EXPORT] Handling old schema data mapping")
            # For old schema, the important columns are already in Vietnamese after rename
            # Just ensure we have the right data
            
            # Don't remove the essential old columns yet - just remove true duplicates
            columns_to_remove = []
            
            # Check for exact duplicates where we have both old and new versions
            if 'Họ và tên' in df_export.columns and 'full_name' in df_export.columns:
                # If we have both, prefer the one with more data
                if df_export['Họ và tên'].notna().sum() >= df_export['full_name'].notna().sum():
                    columns_to_remove.append('full_name')
                else:
                    columns_to_remove.append('Họ và tên')
                    
            # Remove only true duplicates
            for col in columns_to_remove:
                if col in df_export.columns:
                    df_export = df_export.drop(columns=[col])
                    print(f"[EXPORT] Removed duplicate column: {col}")
        else:
            print("[EXPORT] Using new schema - no duplicate handling needed")
        
        print(f"[EXPORT] Final DataFrame has {len(df_export.columns)} columns")
        
        # Reorder columns for proper display order - use Vietnamese column names after mapping
        order_vietnamese = [
            'STT', 'Email', 'Họ và tên', 'Tên gọi khác', 'Lớp', 'Ngày sinh', 'Giới tính', 
            'Dân tộc', 'Quốc tịch', 'Tôn giáo', 'Số điện thoại', 'Số CCCD', 'Ngày cấp CCCD', 'Nơi cấp CCCD', 'Mã định danh', 
            'Số hộ chiếu', 'Ngày cấp hộ chiếu', 'Nơi cấp hộ chiếu', 'Đoàn/Đội',
            'Tỉnh thường trú', 'Phường thường trú', 'Khu phố thường trú', 'Địa chỉ thường trú',
            'Tỉnh quê quán', 'Phường quê quán', 'Khu phố quê quán',
            'Tỉnh cấp giấy khai sinh', 'Phường cấp giấy khai sinh', 'Tỉnh nơi sinh', 'Phường nơi sinh',
            'Địa chỉ chi tiết hiện tại', 'Tỉnh hiện tại', 'Phường hiện tại', 'Khu phố hiện tại',
            'Chiều cao (cm)', 'Cân nặng (kg)', 'Tật khúc xạ (mắt)', 'Kỹ năng bơi',
            'Điện thoại thông minh', 'Máy tính',
            'Họ tên cha', 'Dân tộc của cha', 'Nghề nghiệp cha', 'Năm sinh cha', 'SĐT cha', 'CCCD cha',
            'Họ tên mẹ', 'Dân tộc của mẹ', 'Nghề nghiệp mẹ', 'Năm sinh mẹ', 'SĐT mẹ', 'CCCD mẹ',
            'Họ tên người giám hộ', 'Nghề nghiệp người giám hộ', 'Năm sinh người giám hộ', 'SĐT người giám hộ', 
            'CCCD người giám hộ', 'Giới tính người giám hộ',
            'Thời gian nộp kê khai'
        ]
        
        # Reorder columns based on the Vietnamese names
        ordered_present = [c for c in order_vietnamese if c in df_export.columns]
        others = [c for c in df_export.columns if c not in ordered_present]
        df_export = df_export[ordered_present + others]

        # Hide empty fields if requested
        if hide_empty_fields:
            # Remove columns that are mostly empty
            for col in df_export.columns:
                if col != 'STT':  # Keep ID column
                    non_null_count = df_export[col].notna().sum()
                    if non_null_count / len(df_export) < 0.1:  # Less than 10% filled
                        df_export = df_export.drop(columns=[col])

        # Create Excel with styling
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active
            ws.title = "Danh sách học sinh"

            # Add title if requested
            if include_stats:
                ws.append([title])
                ws.append([f"Tổng số học sinh: {len(df_export)}"])
                if include_timestamp:
                    ws.append([f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
                ws.append([])  # Empty row

            # Add data
            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws.append(r)

            # Apply styling based on theme
            theme_colors = {
                'blue': '1F4E79',      # Xanh dương đậm
                'green': '2E7D32',     # Xanh lá đậm  
                'orange': 'F57500',    # Cam đậm
                'purple': '7B1FA2',    # Tím đậm
                'red': 'D32F2F',       # Đỏ đậm
                'teal': '00796B'       # Xanh ngọc đậm
            }
            
            header_color = theme_colors.get(theme_color, '1F4E79')
            print(f"[EXCEL] Using theme color: {theme_color} -> #{header_color}")  # Debug log
            header_font = Font(bold=True, color="FFFFFF", size=font_size + 1)  # Header slightly larger
            header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Style header row
            header_row = 5 if include_stats else 1
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Style all data cells with center alignment
            data_alignment = Alignment(horizontal="center", vertical="center")
            data_font = Font(size=font_size)  # Use dynamic font size
            
            print(f"[XLSX] Applying font size: {font_size}")  # Debug log
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Apply center alignment to all cells
                    cell.alignment = data_alignment
                    
                    # Apply font to data rows (not header)
                    if row != header_row:
                        cell.font = data_font

            # Auto-resize columns to fit content
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        # Handle Vietnamese characters properly
                        cell_length = len(cell_value)
                        # Add extra padding for Vietnamese characters
                        if any(ord(char) > 127 for char in cell_value):
                            cell_length = int(cell_length * 1.2)  # 20% extra for Vietnamese
                        
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

                # Set minimum width of 12, maximum of 80, with 3 characters padding
                adjusted_width = min(max(max_length + 3, 12), 80)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Set row heights
            for row in range(1, ws.max_row + 1):
                if row == header_row:
                    ws.row_dimensions[row].height = 30  # Header taller
                else:
                    ws.row_dimensions[row].height = 25  # Data rows

            wb.save(filename)

        except ImportError:
            return jsonify({'error': "Thiếu thư viện 'openpyxl'. Vui lòng cài đặt: pip install openpyxl"}), 500
        except Exception as e:
            print(f"[XLSX] Warning: Styling failed, using basic export: {str(e)}")
            df_export.to_excel(filename, index=False)

        return send_file_with_cleanup(filename, as_attachment=True, download_name=filename)

    except Exception as e:
        print(f"[XLSX] Error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-csv', methods=['GET'])
def export_csv():
    """Export to CSV format"""
    try:
        # Get parameters (similar to xlsx but simpler)
        export_type = request.args.get('type', 'all')
        grade = request.args.get('grade')  
        classes = request.args.get('classes')
        province = request.args.get('province')  # Province filter
        ethnicity = request.args.get('ethnicity')  # Ethnicity filter
        
        # Auto detect export type if not provided
        if export_type == 'all':
            if grade:
                export_type = 'grade'
            elif classes:
                export_type = 'class'
            elif (request.args.get('gender') or 
                  request.args.get('fromYear') or 
                  request.args.get('toYear') or 
                  request.args.get('hasPhone') or
                  province or ethnicity):
                export_type = 'custom'
        
        conn = get_db_connection()
        
        # Build query - select only needed columns (same as XLSX)
        basic_columns = [
            'id', 'ho_ten', 'ngay_sinh', 'gioi_tinh', 'lop', 'khoi', 
            'sdt', 'email', 'created_at', 'nickname', 'nationality', 
            'citizen_id', 'cccd_date', 'cccd_place', 'personal_id', 
            'passport', 'passport_date', 'passport_place', 
            'organization', 'permanent_province', 'permanent_ward', 
            'permanent_hamlet', 'permanent_street', 'hometown_province', 
            'hometown_ward', 'hometown_hamlet', 'current_ward', 
            'current_hamlet', 'birthplace_province', 'birthplace_ward', 
            'birth_cert_province', 'birth_cert_ward', 'height', 'weight', 
            'eye_diseases', 'swimming_skill', 'smartphone', 'computer', 
            'father_ethnicity', 'father_birth_year', 'father_phone', 
            'father_cccd', 'mother_ethnicity', 'mother_birth_year', 
            'mother_phone', 'mother_cccd', 'guardian_name', 'guardian_job', 
            'guardian_birth_year', 'guardian_phone', 'guardian_cccd', 
            'guardian_gender'
        ]
        
        # Add filter columns if needed
        if ethnicity:
            if 'dan_toc' not in basic_columns:
                basic_columns.append('dan_toc')
        
        column_list = ', '.join(basic_columns)
        base_query = f'SELECT {column_list} FROM students'
        where_conditions = []
        query_params = []

        if export_type == 'grade' and grade:
            placeholder = get_placeholder()
            if DB_CONFIG['type'] == 'postgres':
                where_conditions.append(f"substring(lop, 1, LENGTH({placeholder})) = {placeholder}")
            else:
                where_conditions.append(f"SUBSTR(lop, 1, LENGTH({placeholder})) = {placeholder}")
            query_params.extend([grade, grade])
        elif export_type == 'class' and classes:
            class_list = [cls.strip() for cls in classes.split(',')]
            placeholder = get_placeholder()
            placeholders = ','.join([placeholder for _ in class_list])
            where_conditions.append(f"lop IN ({placeholders})")
            query_params.extend(class_list)
        elif export_type == 'custom':
            # Handle custom filters for CSV
            gender = request.args.get('gender')
            from_year = request.args.get('fromYear')
            to_year = request.args.get('toYear')
            has_phone = request.args.get('hasPhone') == 'true'
            
            if gender:
                gender_list = [g.strip() for g in gender.split(',')]
                placeholder = get_placeholder()
                gender_placeholders = ','.join([placeholder for _ in gender_list])
                where_conditions.append(f"gioi_tinh IN ({gender_placeholders})")
                query_params.extend(gender_list)
                
            if from_year:
                placeholder = get_placeholder()
                if DB_CONFIG['type'] == 'postgres':
                    where_conditions.append(f"CAST(substring(ngay_sinh, 1, 4) AS INTEGER) >= {placeholder}")
                else:
                    where_conditions.append(f"CAST(SUBSTR(ngay_sinh, 1, 4) AS INTEGER) >= {placeholder}")
                query_params.append(int(from_year))
                
            if to_year:
                placeholder = get_placeholder()
                if DB_CONFIG['type'] == 'postgres':
                    where_conditions.append(f"CAST(substring(ngay_sinh, 1, 4) AS INTEGER) <= {placeholder}")
                else:
                    where_conditions.append(f"CAST(SUBSTR(ngay_sinh, 1, 4) AS INTEGER) <= {placeholder}")
                query_params.append(int(to_year))
                
            if has_phone:
                where_conditions.append("sdt IS NOT NULL AND sdt != ''")

        # Apply province and ethnicity filters for ALL export types
        if province:
            placeholder = get_placeholder()
            # Flexible province matching - case insensitive and partial match
            where_conditions.append(f"LOWER(permanent_province) LIKE LOWER({placeholder})")
            query_params.append(f"%{province}%")
            print(f"[CSV] Filtering by province: %{province}%")

        if ethnicity:
            placeholder = get_placeholder()
            # Flexible ethnicity matching - case insensitive and partial match
            where_conditions.append(f"LOWER(ethnicity) LIKE LOWER({placeholder})")
            query_params.append(f"%{ethnicity}%")
            print(f"[CSV] Filtering by ethnicity: %{ethnicity}%")

        if where_conditions:
            query = f"{base_query} WHERE {' AND '.join(where_conditions)} ORDER BY id ASC"
        else:
            query = f"{base_query} ORDER BY id ASC"

        # Execute query
        if query_params:
            cursor = conn.cursor()
            cursor.execute(query, query_params)
            rows = cursor.fetchall()
            column_names = [description[0] for description in cursor.description]
            df_final = pd.DataFrame(rows, columns=column_names)
        else:
            df_final = pd.read_sql_query(query, conn)

        conn.close()

        if df_final.empty:
            return jsonify({'error': 'Không có dữ liệu phù hợp để xuất'}), 400

        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if export_type == 'grade' and grade:
            filename = f'danh_sach_hoc_sinh_khoi_{grade}_{timestamp}.csv'
        elif export_type == 'class' and classes:
            class_list = [cls.strip() for cls in classes.split(',')]
            if len(class_list) == 1:
                filename = f'danh_sach_hoc_sinh_lop_{class_list[0]}_{timestamp}.csv'
            else:
                filename = f'danh_sach_hoc_sinh_{len(class_list)}_lop_{timestamp}.csv'
        elif export_type == 'custom':
            filename = f'danh_sach_hoc_sinh_tuy_chinh_{timestamp}.csv'
        else:
            filename = f'danh_sach_hoc_sinh_tat_ca_{timestamp}.csv'

        # Column mapping for Vietnamese headers
        column_mapping = {
            'id': 'STT',
            'email': 'Email',
            'full_name': 'Họ và tên',
            'nickname': 'Tên gọi khác',
            'class': 'Lớp',
            'birth_date': 'Ngày sinh',
            'gender': 'Giới tính',
            'ethnicity': 'Dân tộc',
            'nationality': 'Quốc tịch',
            'religion': 'Tôn giáo',
            'phone': 'Số điện thoại',
            'citizen_id': 'Số CCCD',
            'cccd_date': 'Ngày cấp CCCD',
            'cccd_place': 'Nơi cấp CCCD',
            'personal_id': 'Mã định danh',
            'passport': 'Số hộ chiếu',
            'passport_date': 'Ngày cấp hộ chiếu',
            'passport_place': 'Nơi cấp hộ chiếu',
            'organization': 'Đoàn/Đội',
            'permanent_province': 'Tỉnh thường trú',
            'permanent_ward': 'Phường thường trú',
            'permanent_hamlet': 'Khu phố thường trú',
            'permanent_street': 'Địa chỉ thường trú',
            'hometown_province': 'Tỉnh quê quán',
            'hometown_ward': 'Phường quê quán',
            'hometown_hamlet': 'Khu phố quê quán',
            'birth_cert_province': 'Tỉnh cấp giấy khai sinh',
            'birth_cert_ward': 'Phường cấp giấy khai sinh',
            'birthplace_province': 'Tỉnh nơi sinh',
            'birthplace_ward': 'Phường nơi sinh',
            'current_address_detail': 'Địa chỉ hiện tại',
            'current_province': 'Tỉnh hiện tại',
            'current_ward': 'Phường hiện tại',
            'current_hamlet': 'Khu phố hiện tại',
            'height': 'Chiều cao (cm)',
            'weight': 'Cân nặng (kg)',
            'eye_diseases': 'Tật khúc xạ (mắt)',
            'swimming_skill': 'Kỹ năng bơi',
            'smartphone': 'Điện thoại thông minh',
            'computer': 'Máy tính',
            'father_ethnicity': 'Dân tộc của cha',
            'mother_ethnicity': 'Dân tộc của mẹ',
            'father_name': 'Họ tên cha',
            'father_job': 'Nghề nghiệp cha',
            'father_birth_year': 'Năm sinh cha',
            'father_phone': 'SĐT cha',
            'father_cccd': 'CCCD cha',
            'mother_name': 'Họ tên mẹ',
            'mother_job': 'Nghề nghiệp mẹ',
            'mother_birth_year': 'Năm sinh mẹ',
            'mother_phone': 'SĐT mẹ',
            'mother_cccd': 'CCCD mẹ',
            'guardian_name': 'Họ tên người giám hộ',
            'guardian_job': 'Nghề nghiệp người giám hộ',
            'guardian_birth_year': 'Năm sinh người giám hộ',
            'guardian_phone': 'SĐT người giám hộ',
            'guardian_cccd': 'CCCD người giám hộ',
            'guardian_gender': 'Giới tính người giám hộ',
            'created_at': 'Thời gian nộp kê khai'
        }

        # Apply column mapping and reorder columns
        df_export = df_final.rename(columns=column_mapping)
        
        # Ensure created_at column appears at the end
        order_keys = [
            'id',
            'email','full_name','nickname','class','birth_date','gender','ethnicity','nationality','religion','phone',
            'citizen_id','cccd_date','cccd_place','personal_id','passport','passport_date','passport_place','organization',
            'permanent_province','permanent_ward','permanent_hamlet','permanent_street',
            'hometown_province','hometown_ward','hometown_hamlet',
            'birth_cert_province','birth_cert_ward','birthplace_province','birthplace_ward',
            'current_address_detail','current_province','current_ward','current_hamlet',
            'height','weight','eye_diseases','swimming_skill',
            'smartphone','computer',
            'father_name','father_ethnicity','father_job','father_birth_year','father_phone','father_cccd',
            'mother_name','mother_ethnicity','mother_job','mother_birth_year','mother_phone','mother_cccd',
            'guardian_name','guardian_job','guardian_birth_year','guardian_phone','guardian_cccd','guardian_gender',
            'created_at'
        ]
        order_vn = [column_mapping.get(k, k) for k in order_keys]
        ordered_present = [c for c in order_vn if c in df_export.columns]
        others = [c for c in df_export.columns if c not in ordered_present]
        df_export = df_export[ordered_present + others]

        # Export to CSV
        df_export.to_csv(filename, index=False, encoding='utf-8-sig')
        
        return send_file_with_cleanup(filename, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-pdf', methods=['GET'])
def export_pdf():
    """Export to PDF format"""
    try:
        return jsonify({'error': 'PDF export đang được phát triển'}), 501
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-json', methods=['GET'])
def export_json():
    """Export to JSON format"""
    try:
        # Get parameters
        export_type = request.args.get('type', 'all')
        grade = request.args.get('grade')  
        classes = request.args.get('classes')
        province = request.args.get('province')  # Province filter
        ethnicity = request.args.get('ethnicity')  # Ethnicity filter
        
        # Auto detect export type if not provided
        if export_type == 'all':
            if grade:
                export_type = 'grade'
            elif classes:
                export_type = 'class'
            elif (request.args.get('gender') or 
                  request.args.get('fromYear') or 
                  request.args.get('toYear') or 
                  request.args.get('hasPhone') or
                  province or ethnicity):
                export_type = 'custom'
        
        conn = get_db_connection()
        
        # Build query
        base_query = 'SELECT * FROM students'
        where_conditions = []
        query_params = []

        if export_type == 'grade' and grade:
            if DB_CONFIG['type'] == 'postgres':
                where_conditions.append("substring(class, 1, LENGTH(?)) = ?")
            else:
                where_conditions.append("SUBSTR(class, 1, LENGTH(?)) = ?")
            query_params.extend([grade, grade])
        elif export_type == 'class' and classes:
            class_list = [cls.strip() for cls in classes.split(',')]
            placeholders = ','.join(['?' for _ in class_list])
            where_conditions.append(f"class IN ({placeholders})")
            query_params.extend(class_list)
        elif export_type == 'custom':
            # Handle custom filters for JSON
            gender = request.args.get('gender')
            from_year = request.args.get('fromYear')
            to_year = request.args.get('toYear')
            has_phone = request.args.get('hasPhone') == 'true'
            
            if gender:
                gender_list = [g.strip() for g in gender.split(',')]
                gender_placeholders = ','.join(['?' for _ in gender_list])
                where_conditions.append(f"gender IN ({gender_placeholders})")
                query_params.extend(gender_list)
                
            if from_year:
                if DB_CONFIG['type'] == 'postgres':
                    where_conditions.append("CAST(substring(birth_date, 1, 4) AS INTEGER) >= ?")
                else:
                    where_conditions.append("CAST(SUBSTR(birth_date, 1, 4) AS INTEGER) >= ?")
                query_params.append(int(from_year))
                
            if to_year:
                if DB_CONFIG['type'] == 'postgres':
                    where_conditions.append("CAST(substring(birth_date, 1, 4) AS INTEGER) <= ?")
                else:
                    where_conditions.append("CAST(SUBSTR(birth_date, 1, 4) AS INTEGER) <= ?")
                query_params.append(int(to_year))
                
            if has_phone:
                where_conditions.append("phone IS NOT NULL AND phone != ''")

        # Apply province and ethnicity filters for ALL export types
        if province:
            where_conditions.append("permanent_province = ?")
            query_params.append(province)
            print(f"[JSON] Filtering by province: {province}")

        if ethnicity:
            where_conditions.append("ethnicity = ?")
            query_params.append(ethnicity)
            print(f"[JSON] Filtering by ethnicity: {ethnicity}")

        if where_conditions:
            query = f"{base_query} WHERE {' AND '.join(where_conditions)} ORDER BY id ASC"
        else:
            query = f"{base_query} ORDER BY id ASC"

        # Execute query
        if query_params:
            cursor = conn.cursor()
            cursor.execute(query, query_params)
            rows = cursor.fetchall()
            column_names = [description[0] for description in cursor.description]
            df_final = pd.DataFrame(rows, columns=column_names)
        else:
            df_final = pd.read_sql_query(query, conn)

        conn.close()

        if df_final.empty:
            return jsonify({'error': 'Không có dữ liệu phù hợp để xuất'}), 400

        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if export_type == 'grade' and grade:
            filename = f'danh_sach_hoc_sinh_khoi_{grade}_{timestamp}.json'
        elif export_type == 'class' and classes:
            class_list = [cls.strip() for cls in classes.split(',')]
            if len(class_list) == 1:
                filename = f'danh_sach_hoc_sinh_lop_{class_list[0]}_{timestamp}.json'
            else:
                filename = f'danh_sach_hoc_sinh_{len(class_list)}_lop_{timestamp}.json'
        elif export_type == 'custom':
            filename = f'danh_sach_hoc_sinh_tuy_chinh_{timestamp}.json'
        else:
            filename = f'danh_sach_hoc_sinh_tat_ca_{timestamp}.json'

        # Column mapping for Vietnamese headers
        column_mapping = {
            'id': 'STT',
            'email': 'Email',
            'full_name': 'Họ và tên',
            'nickname': 'Tên gọi khác',
            'class': 'Lớp',
            'birth_date': 'Ngày sinh',
            'gender': 'Giới tính',
            'ethnicity': 'Dân tộc',
            'nationality': 'Quốc tịch',
            'religion': 'Tôn giáo',
            'phone': 'Số điện thoại',
            'citizen_id': 'Số CCCD',
            'cccd_date': 'Ngày cấp CCCD',
            'cccd_place': 'Nơi cấp CCCD',
            'personal_id': 'Mã định danh',
            'passport': 'Số hộ chiếu',
            'passport_date': 'Ngày cấp hộ chiếu',
            'passport_place': 'Nơi cấp hộ chiếu',
            'organization': 'Đoàn/Đội',
            'permanent_province': 'Tỉnh thường trú',
            'permanent_ward': 'Phường thường trú',
            'permanent_hamlet': 'Khu phố thường trú',
            'permanent_street': 'Địa chỉ thường trú',
            'hometown_province': 'Tỉnh quê quán',
            'hometown_ward': 'Phường quê quán',
            'hometown_hamlet': 'Khu phố quê quán',
            'birth_cert_province': 'Tỉnh cấp giấy khai sinh',
            'birth_cert_ward': 'Phường cấp giấy khai sinh',
            'birthplace_province': 'Tỉnh nơi sinh',
            'birthplace_ward': 'Phường nơi sinh',
            'current_address_detail': 'Địa chỉ hiện tại',
            'current_province': 'Tỉnh hiện tại',
            'current_ward': 'Phường hiện tại',
            'current_hamlet': 'Khu phố hiện tại',
            'height': 'Chiều cao (cm)',
            'weight': 'Cân nặng (kg)',
            'eye_diseases': 'Tật khúc xạ (mắt)',
            'swimming_skill': 'Kỹ năng bơi',
            'smartphone': 'Điện thoại thông minh',
            'computer': 'Máy tính',
            'father_ethnicity': 'Dân tộc của cha',
            'mother_ethnicity': 'Dân tộc của mẹ',
            'father_name': 'Họ tên cha',
            'father_job': 'Nghề nghiệp cha',
            'father_birth_year': 'Năm sinh cha',
            'father_phone': 'SĐT cha',
            'father_cccd': 'CCCD cha',
            'mother_name': 'Họ tên mẹ',
            'mother_job': 'Nghề nghiệp mẹ',
            'mother_birth_year': 'Năm sinh mẹ',
            'mother_phone': 'SĐT mẹ',
            'mother_cccd': 'CCCD mẹ',
            'guardian_name': 'Họ tên người giám hộ',
            'guardian_job': 'Nghề nghiệp người giám hộ',
            'guardian_birth_year': 'Năm sinh người giám hộ',
            'guardian_phone': 'SĐT người giám hộ',
            'guardian_cccd': 'CCCD người giám hộ',
            'guardian_gender': 'Giới tính người giám hộ',
            'created_at': 'Thời gian nộp kê khai'
        }

        # Apply column mapping and reorder columns
        df_export = df_final.rename(columns=column_mapping)
        
        # Ensure created_at column appears at the end
        order_keys = [
            'id',
            'email','full_name','nickname','class','birth_date','gender','ethnicity','nationality','religion','phone',
            'citizen_id','cccd_date','cccd_place','personal_id','passport','passport_date','passport_place','organization',
            'permanent_province','permanent_ward','permanent_hamlet','permanent_street',
            'hometown_province','hometown_ward','hometown_hamlet',
            'birth_cert_province','birth_cert_ward','birthplace_province','birthplace_ward',
            'current_address_detail','current_province','current_ward','current_hamlet',
            'height','weight','eye_diseases','swimming_skill',
            'smartphone','computer',
            'father_name','father_ethnicity','father_job','father_birth_year','father_phone','father_cccd',
            'mother_name','mother_ethnicity','mother_job','mother_birth_year','mother_phone','mother_cccd',
            'guardian_name','guardian_job','guardian_birth_year','guardian_phone','guardian_cccd','guardian_gender',
            'created_at'
        ]
        order_vn = [column_mapping.get(k, k) for k in order_keys]
        ordered_present = [c for c in order_vn if c in df_export.columns]
        others = [c for c in df_export.columns if c not in ordered_present]
        df_export = df_export[ordered_present + others]

        # Export to JSON
        result = {
            "export_info": {
                "title": "Danh sách học sinh THPT Dĩ An",
                "exported_at": datetime.now().isoformat(),
                "total_records": len(df_export),
                "export_type": export_type
            },
            "data": df_export.to_dict('records')
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        return send_file_with_cleanup(filename, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


LOCATIONS_LATEST = None
LOCATIONS_SOURCE = 'none'

def _first_existing_column(df, candidates):
    cols = [c.strip() if isinstance(c, str) else c for c in df.columns]
    df.columns = cols
    for cand in candidates:
        if cand in df.columns:
            return cand
    return None

def _is_nan(v):
    try:
        return v is None or (isinstance(v, float) and math.isnan(v)) or (isinstance(v, str) and not v.strip())
    except Exception:
        return v is None

def load_locations_latest():
    global LOCATIONS_LATEST, LOCATIONS_SOURCE
    if LOCATIONS_LATEST is not None:
        return LOCATIONS_LATEST

    base_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(base_dir, 'final_danh-muc-phuong-xa_moi.xlsx')
    csv_path = os.path.join(base_dir, 'final_danh-muc-phuong-xa_moi.csv')

    df = None
    try:
        if os.path.exists(xlsx_path):
            df = pd.read_excel(xlsx_path, engine='openpyxl')
            LOCATIONS_SOURCE = 'xlsx'
    except Exception:
        df = None
        LOCATIONS_SOURCE = 'none'
    if df is None:
        try:
            if os.path.exists(csv_path):
                for header_row in [2, 1, 0]:
                    try:
                        df = pd.read_csv(
                            csv_path,
                            header=header_row,
                            encoding='utf-8-sig',
                            keep_default_na=False
                        )
                        if df is not None and len(df.columns) >= 6:
                            LOCATIONS_SOURCE = 'csv'
                            break
                    except Exception:
                        continue
        except Exception:
            df = None
            LOCATIONS_SOURCE = 'none'

    if df is None or df.empty:
        LOCATIONS_LATEST = {'provinces': [], 'wardsByProvince': {}, 'meta': {'source': LOCATIONS_SOURCE, 'provinces': 0, 'wardKeys': 0}}
        print('[LOC] Catalog empty or not loaded. source=', LOCATIONS_SOURCE)
        return LOCATIONS_LATEST

    province_code_col = _first_existing_column(df, [
        'Mã tỉnh (BNV)', 'Ma tinh (BNV)', 'Mã tỉnh', 'Mã Tỉnh (BNV)'
    ])
    province_name_col = _first_existing_column(df, [
        'Tên tỉnh/TP mới', 'Ten tinh/TP moi', 'Tên Tỉnh/TP mới', 'Tên tỉnh / TP mới'
    ])
    district_code_col = _first_existing_column(df, [
        'Mã Quận huyện TMS (cũ) CQT đã rà soát', 'Mã Quận huyện TMS (cũ)', 'Mã quận huyện', 'Ma Quan huyen TMS (cu)'
    ])
    district_name_col = _first_existing_column(df, [
        'Tên Quận huyện TMS (cũ)', 'Tên quận huyện', 'Ten Quan huyen TMS (cu)'
    ])
    ward_code_col = _first_existing_column(df, [
        'Mã phường/xã mới', 'Mã phường/xã mới ', 'Ma phuong/xa moi', 'Mã Phường/Xã mới'
    ])
    ward_name_col = _first_existing_column(df, [
        'Tên Phường/Xã mới', 'Ten Phuong/Xa moi', 'Tên phường/xã mới'
    ])

    if not all([province_code_col, province_name_col, district_code_col, district_name_col, ward_code_col, ward_name_col]):
        try:
            cols = list(df.columns)
            if len(cols) >= 10:
                province_code_col = province_code_col or cols[2]
                province_name_col = province_name_col or cols[3]
                district_code_col = district_code_col or cols[5]
                district_name_col = district_name_col or cols[6]
                ward_code_col = ward_code_col or cols[8]
                ward_name_col = ward_name_col or cols[9]
                print('[LOC] Using index-based header fallback')
            else:
                LOCATIONS_LATEST = {'provinces': [], 'wardsByProvince': {}, 'meta': {'source': LOCATIONS_SOURCE, 'provinces': 0, 'wardKeys': 0}}
                print('[LOC] Columns insufficient:', len(cols))
                return LOCATIONS_LATEST
        except Exception as e:
            LOCATIONS_LATEST = {'provinces': [], 'wardsByProvince': {}, 'meta': {'source': LOCATIONS_SOURCE, 'provinces': 0, 'wardKeys': 0, 'error': str(e)}}
            print('[LOC] Header detection failed:', e)
            return LOCATIONS_LATEST

    provinces_map = {}
    wards_by_province = {}  # Thay đổi: bỏ districts, chỉ có wards theo province

    for _, row in df.iterrows():
        pc = row.get(province_code_col)
        pn = row.get(province_name_col)
        dc = row.get(district_code_col)
        dn = row.get(district_name_col)
        wc = row.get(ward_code_col)
        wn = row.get(ward_name_col)

        if _is_nan(pc) or _is_nan(pn) or _is_nan(wc) or _is_nan(wn):
            continue

        pc = str(pc).strip()
        pn = str(pn).strip()
        wc = str(wc).strip()
        wn = str(wn).strip()

        # Bỏ qua dòng header hoặc dữ liệu không hợp lệ
        if (pn in ['Tên tỉnh/TP mới', 'Ten tinh/TP moi', 'Tên Tỉnh/TP mới', 'Tên tỉnh / TP mới'] or
            wn in ['Tên Phường/Xã mới', 'Ten Phuong/Xa moi', 'Tên phường/xã mới'] or
            pc == 'Mã tỉnh (BNV)' or wc == 'Mã phường/xã mới'):
            continue

        # Chuẩn hóa tên thành phố từ "Tp" thành "Thành phố"
        if pn.startswith('Tp '):
            pn = pn.replace('Tp ', 'Thành phố ')

        if pc not in provinces_map:
            provinces_map[pc] = pn
        if pc not in wards_by_province:
            wards_by_province[pc] = []
        wards_by_province[pc].append({'code': wc, 'name': wn})

    provinces = [{'code': code, 'name': name} for code, name in provinces_map.items()]
    provinces.sort(key=lambda x: x['name'])

    wards_by_province_sorted = {}
    for pc, wards in wards_by_province.items():
        wards_sorted = sorted(wards, key=lambda x: x['name'])
        wards_by_province_sorted[pc] = wards_sorted

    LOCATIONS_LATEST = {
        'provinces': provinces,
        'wardsByProvince': wards_by_province_sorted,  # Thay đổi: chỉ có wards theo province
        'meta': {
            'source': LOCATIONS_SOURCE,
            'provinces': len(provinces),
            'wardKeys': len(wards_by_province_sorted.keys())
        }
    }
    try:
        print(f"[LOC] Loaded: src={LOCATIONS_SOURCE}, provinces={len(provinces)}, wardKeys={len(wards_by_province_sorted.keys())}")
    except Exception:
        pass
    return LOCATIONS_LATEST

@app.route('/api/locations/latest', methods=['GET'])
def api_locations_latest():
    try:
        refresh = request.args.get('refresh')
        global LOCATIONS_LATEST
        if refresh in ('1', 'true', 'yes'):
            LOCATIONS_LATEST = None
        data = load_locations_latest()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete-student/<int:student_id>', methods=['DELETE'])
def delete_student(student_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(convert_placeholders('DELETE FROM students WHERE id = ?'), (student_id,))

        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Không tìm thấy học sinh'}), 404

        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': 'Đã xóa học sinh thành công'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin-login', methods=['POST'])
def admin_login_api():
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        password = data.get('password', '')

        if not email or not password:
            return jsonify({'error': 'Vui lòng nhập đầy đủ thông tin'}), 400

        if email not in ADMIN_ACCOUNTS or ADMIN_ACCOUNTS[email] != password:
            return jsonify({'error': 'Email hoặc mật khẩu không đúng'}), 401

        otp = generate_otp()
        store_otp(email, otp)

        email_sent = False
        if not FORCE_CONSOLE_OTP:
            email_sent = send_otp_email(email, otp)

        if email_sent:
            return jsonify({
                'success': True,
                'message': 'Mã OTP đã được gửi đến email của bạn. Kiểm tra cả thư mục spam.',
                'email': email,
                'fallback': False
            })
        else:
            return jsonify({
                'success': True,
                'message': f'Mã OTP được hiển thị trong console và giao diện',
                'email': email,
                'fallback': True,
                'debug_otp': otp
            })

    except Exception as e:
        print(f"[LOGIN ERROR] {str(e)}")
        return jsonify({'error': 'Có lỗi xảy ra khi đăng nhập'}), 500

@app.route('/api/verify-otp', methods=['POST'])
def verify_otp_api():
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        otp = data.get('otp', '').strip()

        if not email or not otp:
            return jsonify({'error': 'Vui lòng nhập đầy đủ thông tin'}), 400

        success, message = verify_otp(email, otp)

        if success:
            return jsonify({
                'success': True,
                'message': message,
                'session': {
                    'email': email,
                    'loginTime': int(time.time() * 1000),
                    'expiry': int((time.time() + 24 * 60 * 60) * 1000)
                }
            })
        else:
            return jsonify({'error': message}), 400

    except Exception as e:
        print(f"[OTP ERROR] {str(e)}")
        return jsonify({'error': 'Có lỗi xảy ra khi xác thực OTP'}), 500

@app.route('/api/resend-otp', methods=['POST'])
def resend_otp_api():
    try:
        data = request.get_json()
        email = data.get('email', '').strip()

        if not email:
            return jsonify({'error': 'Email không hợp lệ'}), 400

        if email not in ADMIN_ACCOUNTS:
            return jsonify({'error': 'Email không được phép'}), 401

        otp = generate_otp()
        store_otp(email, otp)

        email_sent = False
        if not FORCE_CONSOLE_OTP:
            email_sent = send_otp_email(email, otp)

        if email_sent:
            return jsonify({
                'success': True,
                'message': 'Mã OTP mới đã được gửi đến email của bạn. Kiểm tra cả thư mục spam.',
                'fallback': False
            })
        else:
            return jsonify({
                'success': True,
                'message': f'Mã OTP mới được hiển thị trong console và giao diện',
                'fallback': True,
                'debug_otp': otp
            })

    except Exception as e:
        print(f"[RESEND ERROR] {str(e)}")
        return jsonify({'error': 'Có lỗi xảy ra khi gửi lại OTP'}), 500

@app.route('/admin')
def admin():
    return send_file('admin-login.html')

@app.route('/admin-login.html')
def admin_login():
    return send_file('admin-login.html')

@app.route('/admin-panel')
def admin_panel():
    return send_file('admin.html')

@app.route('/admin.html')
def admin_html():
    return send_file('admin.html')

if __name__ == '__main__':
    init_db()
    migrate_db()

    email_working = test_email_config()

    try:
        print(" Server đang chạy tại: http://localhost:5000")
        print(" Trang admin: http://localhost:5000/admin")
        print(f" Số tài khoản admin: {len(ADMIN_ACCOUNTS)}")
        
        if SHOW_ADMIN_CREDENTIALS:
            print("Tài khoản admin:")
            for email, password in ADMIN_ACCOUNTS.items():
                print(f"   • {email} / {password}")
        else:
            print("Xem file .env để biết thông tin đăng nhập")
            
        if email_working:
            print("Email OTP: Đã cấu hình và hoạt động")
        else:
            print("⚠️  Email OTP: Lỗi cấu hình, sử dụng console debug")
    except Exception:
        print("Server dang chay tai: http://localhost:5000")
        print("Trang admin: http://localhost:5000/admin")
        print(f"So tai khoan admin: {len(ADMIN_ACCOUNTS)}")

# Admin bulk operations endpoints
@app.route('/api/clear-all-data', methods=['DELETE'])
def clear_all_data():
    """Xóa tất cả học sinh và reset auto increment (như clear_data.py)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM students")
        count = cursor.fetchone()[0]
        
        # Xóa tất cả dữ liệu và reset auto increment
        cursor.execute("DELETE FROM students")
        # Reset auto increment dựa vào database type
        if DB_CONFIG['type'] == 'postgresql':
            cursor.execute("ALTER SEQUENCE students_id_seq RESTART WITH 1")
        else:
            cursor.execute("DELETE FROM sqlite_sequence WHERE name='students'")
        conn.commit()
        conn.close()
        
        print(f"[ADMIN] Đã xóa tất cả {count} học sinh và reset database")
        return jsonify({'success': True, 'deleted_count': count})
    except Exception as e:
        print(f"[ERROR] Clear all data: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generate-sample-data', methods=['POST'])
def generate_sample_data():
    """Tạo dữ liệu mẫu thực tế (như generate_sample_data.py)"""
    try:
        print("[DEBUG] Starting generate sample data...")
        
        # Handle JSON request safely
        if request.is_json:
            data = request.get_json() or {}
        else:
            data = {}
            
        count = int(data.get('count', 50))
        print(f"[DEBUG] Count requested: {count}")
        
        if count > 200:
            return jsonify({'success': False, 'error': 'Không thể tạo quá 200 bản ghi cùng lúc'}), 400
        
        if count <= 0:
            return jsonify({'success': False, 'error': 'Số lượng phải lớn hơn 0'}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        print("[DEBUG] Database connection established")
        
        # Check existing columns first to avoid column errors
        if DB_CONFIG['type'] == 'postgresql':
            cursor.execute("""
                SELECT column_name FROM information_schema.columns 
                WHERE table_name = 'students'
            """)
            existing_columns = [row[0] for row in cursor.fetchall()]
        else:
            cursor.execute('PRAGMA table_info(students)')
            existing_columns = [col[1] for col in cursor.fetchall()]
        
        print(f"[DEBUG] Found {len(existing_columns)} columns in database")
        
        import random
        from datetime import datetime, timedelta
        
        # Dữ liệu mẫu thực tế từ generate_sample_data.py
        first_names = [
            'Nguyễn', 'Trần', 'Lê', 'Phạm', 'Hoàng', 'Huỳnh', 'Phan', 'Vũ', 'Võ', 'Đặng',
            'Bùi', 'Đỗ', 'Hồ', 'Ngô', 'Dương', 'Lý', 'Đinh', 'Đào', 'Cao', 'Lương', 'Mai'
        ]
        middle_names = ['Văn', 'Thị', 'Minh', 'Hoàng', 'Quang', 'Hữu', 'Thanh', 'Anh', 'Thành', 'Bảo', 'Vinh', 'Khánh', 'Nhật', 'Mai', 'Ngọc']
        last_names = [
            'An', 'Bình', 'Cường', 'Dũng', 'Đức', 'Giang', 'Hà', 'Hải', 'Khang', 'Linh',
            'Long', 'Mai', 'Minh', 'Nam', 'Phong', 'Quân', 'Sơn', 'Thảo', 'Tú', 'Vy',
            'Yến', 'Hương', 'Loan', 'Nga', 'Oanh', 'Phương', 'Quyên', 'Thu', 'Trang', 'Xuân'
        ]
        classes = [
            # Khối 10
            '10A1', '10A2', '10A3', '10A4', '10A5', '10A6', '10A7', '10A8',
            '10B1', '10B2', '10B3', '10B4',
            # Khối 11
            '11A1', '11A2', '11A3', '11A4', '11A5', '11A6', '11A7', '11A8',
            '11B1', '11B2', '11B3', '11B4',
            # Khối 12
            '12A1', '12A2', '12A3', '12A4', '12A5', '12A6', '12A7', '12A8',
            '12B1', '12B2', '12B3', '12B4'
        ]
        genders = ['Nam', 'Nữ']
        provinces = ['Thành phố Hồ Chí Minh', 'Tỉnh Đồng Nai', 'Tỉnh Bình Dương', 'Tỉnh Long An', 'Tỉnh Tây Ninh']
        
        created_count = 0
        for i in range(count):
            # Tạo tên thực tế
            first_name = random.choice(first_names)
            middle_name = random.choice(middle_names)
            last_name = random.choice(last_names)
            full_name = f"{first_name} {middle_name} {last_name}"
            
            # Email học sinh mẫu (ảo) - có identifier đặc biệt
            email_name = f"sample_{last_name.lower()}.{middle_name.lower()}.{i+100:03d}_sample"
            email = f"{email_name}@test.sample.com"
            
            # Ngày sinh thực tế (2005-2008)
            birth_year = random.randint(2005, 2008)
            birth_month = random.randint(1, 12)
            birth_day = random.randint(1, 28)
            birth_date = f"{birth_year}-{birth_month:02d}-{birth_day:02d}"
            
            gender = random.choice(genders)
            phone = f"0{random.randint(900000000, 999999999)}"  # SĐT bắt đầu từ 09xx để realistic
            class_name = random.choice(classes)
            province = random.choice(provinces)
            ward = f"Phường {random.randint(1, 20)}"
            street_num = random.randint(1, 500)
            current_address = f"{street_num} Đường {random.randint(1, 50)}, {ward}, {province}"
            
            # CCCD học sinh (12 số)
            cccd_prefix = random.choice(['001', '002', '025', '026', '079'])  # Mã tỉnh thành thực tế
            cccd_number = f"{cccd_prefix}{random.randint(100000000, 999999999)}"
            cccd_date = f"{random.randint(2020, 2024)}-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}"
            cccd_place = f"Công an {province}"
            
            # Thông tin ba mẹ realistic
            father_first = random.choice(first_names)
            father_middle = random.choice(['Văn', 'Minh', 'Thanh', 'Hữu', 'Quang'])
            father_last = random.choice(['An', 'Bình', 'Cường', 'Dũng', 'Hùng', 'Nam', 'Sơn'])
            father_name = f"{father_first} {father_middle} {father_last}"
            father_job = random.choice(['Công nhân', 'Nông dân', 'Giáo viên', 'Bác sĩ', 'Kỹ sư', 'Kinh doanh', 'Công chức', 'Tài xế', 'Thợ xây'])
            father_birth_year = str(random.randint(1970, 1985))
            father_phone = f"0{random.randint(900000000, 999999999)}"
            father_cccd = f"{cccd_prefix}{random.randint(100000000, 999999999)}"
            
            mother_first = random.choice(first_names)
            mother_last_names = ['Lan', 'Hoa', 'Mai', 'Hương', 'Phương', 'Linh', 'Nga', 'Oanh', 'Thu', 'Xuân']
            mother_name = f"{mother_first} Thị {random.choice(mother_last_names)}"
            mother_job = random.choice(['Nội trợ', 'Giáo viên', 'Y tá', 'Kế toán', 'Bán hàng', 'Công nhân', 'Nông dân', 'Nhân viên văn phòng'])
            mother_birth_year = str(random.randint(1975, 1990))
            mother_phone = f"0{random.randint(900000000, 999999999)}"
            mother_cccd = f"{cccd_prefix}{random.randint(100000000, 999999999)}"
            
            # Thời gian tạo ngẫu nhiên trong 30 ngày qua
            days_ago = random.randint(0, 30)
            created_at = (datetime.now() - timedelta(days=days_ago, hours=random.randint(0, 23), minutes=random.randint(0, 59))).isoformat()
            
            # Tạo thêm nhiều thông tin chi tiết để đạt 70% thông tin đầy đủ
            ethnicities = ['Kinh', 'Tày', 'Thái', 'Hoa', 'Mường', 'Nùng', 'H\'Mông', 'Dao', 'Gia Rai', 'Ê Đê']
            religions = ['Không', 'Phật giáo', 'Công giáo', 'Cao Đài', 'Hòa Hảo', 'Tin Lành']
            jobs = [
                'Công nhân', 'Nông dân', 'Giáo viên', 'Bác sĩ', 'Kỹ sư', 'Kinh doanh', 
                'Công chức', 'Tài xế', 'Thợ xây', 'Bán hàng', 'Y tá', 'Kế toán', 
                'Nhân viên văn phòng', 'Nội trợ', 'Lao động tự do'
            ]
            districts = [
                'Huyện Dĩ An', 'Huyện Thuận An', 'Huyện Bến Cát', 'Huyện Tân Uyên',
                'Quận 1', 'Quận 2', 'Quận 3', 'Quận Bình Thạnh', 'Quận Tân Bình',
                'Huyện Nhà Bè', 'Huyện Củ Chi', 'Huyện Hóc Môn'
            ]
            
            current_district = random.choice(districts)
            current_commune = f"Xã {random.choice(['An Phú', 'Tân Hiệp', 'Dĩ An', 'Bình An', 'Tân Thông'])}"
            current_hamlet_detail = f"Ấp {random.randint(1, 5)}"
            
            guardian_gender = random.choice(['Nam', 'Nữ'])
            guardian_relationship = random.choice(['Ông', 'Bà', 'Chú', 'Cô', 'Anh', 'Chị'])
            guardian_first = random.choice(first_names)
            guardian_last = random.choice(last_names)
            guardian_name = f"{guardian_first} {guardian_last}"
            
            # Tạo data với ít nhất 70% thông tin được điền
            all_student_data = {
                # THÔNG TIN CÁ NHÂN (100% điền)
                'email': email,
                'full_name': full_name,
                'birth_date': birth_date,
                'gender': gender,
                'phone': phone,
                'class': class_name,
                'nationality': 'Việt Nam',
                'ethnicity': random.choice(ethnicities),
                'religion': random.choice(religions),
                'nickname': f"{last_name} {random.choice(['nhỏ', 'bé', 'con', 'em', 'tí', 'út'])}",
                'created_at': created_at,
                
                # GIẤY TỜ PHÁP LÝ (100% điền)
                'citizen_id': cccd_number,
                'cccd_date': cccd_date,
                'cccd_place': cccd_place,
                'personal_id': f"HS{random.randint(100000, 999999)}",
                'passport': f"C{random.randint(1000000, 9999999)}" if random.randint(1, 10) <= 3 else None,  # 30% có passport
                'passport_date': f"{random.randint(2020, 2024)}-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}" if random.randint(1, 10) <= 3 else None,
                'passport_place': f"Cục Quản lý xuất nhập cảnh {province}" if random.randint(1, 10) <= 3 else None,
                
                # ĐỊA CHỈ THƯỜNG TRÚ (100% điền)
                'permanent_province': province,
                'permanent_district': current_district,
                'permanent_ward': ward,
                'permanent_hamlet': f"Khu phố {random.randint(1, 10)}",
                'permanent_street': f"{street_num} Đường {random.randint(1, 50)}",
                
                # ĐỊA CHỈ QUÊ QUÁN (90% điền)
                'hometown_province': province if random.randint(1, 10) <= 9 else None,
                'hometown_district': current_district if random.randint(1, 10) <= 9 else None,
                'hometown_ward': ward if random.randint(1, 10) <= 9 else None,
                'hometown_hamlet': f"Khu phố {random.randint(1, 10)}" if random.randint(1, 10) <= 9 else None,
                
                # ĐỊA CHỈ HIỆN TẠI (85% điền)
                'current_province': province if random.randint(1, 10) <= 8 else None,
                'current_district': current_district if random.randint(1, 10) <= 8 else None,
                'current_ward': ward if random.randint(1, 10) <= 8 else None,
                'current_hamlet': f"Khu phố {random.randint(1, 10)}" if random.randint(1, 10) <= 8 else None,
                'current_address_detail': current_address if random.randint(1, 10) <= 8 else None,
                
                # NơI SINH (90% điền)
                'birthplace_province': province if random.randint(1, 10) <= 9 else None,
                'birthplace_district': current_district if random.randint(1, 10) <= 9 else None,
                'birthplace_ward': ward if random.randint(1, 10) <= 9 else None,
                
                # GIẤY KHAI SINH (85% điền)
                'birth_cert_province': province if random.randint(1, 10) <= 8 else None,
                'birth_cert_district': current_district if random.randint(1, 10) <= 8 else None,
                'birth_cert_ward': ward if random.randint(1, 10) <= 8 else None,
                
                # SỨC KHỎE (100% điền)
                'height': random.randint(150, 180),
                'weight': random.randint(45, 75),
                'eye_diseases': random.choice(['Không', 'Cận thị nhẹ', 'Viễn thị nhẹ', 'Loạn thị nhẹ', 'Cận thị nặng']),
                'swimming_skill': random.choice(['Biết bơi', 'Không biết bơi', 'Bơi được 25m', 'Bơi giỏi', 'Bơi cơ bản']),
                
                # THIẾT BỊ HỌC TẬP (100% điền)
                'smartphone': random.choice(['Có', 'Không']),
                'computer': random.choice(['Có', 'Không']),
                
                # THÔNG TIN CHA (100% điền các trường chính)
                'father_name': father_name,
                'father_job': father_job,
                'father_birth_year': father_birth_year,
                'father_phone': father_phone,
                'father_cccd': father_cccd,
                'father_ethnicity': random.choice(ethnicities),
                
                # THÔNG TIN MẸ (100% điền các trường chính)
                'mother_name': mother_name,
                'mother_job': mother_job,
                'mother_birth_year': mother_birth_year,
                'mother_phone': mother_phone,
                'mother_cccd': mother_cccd,
                'mother_ethnicity': random.choice(ethnicities),
                
                # THÔNG TIN NGƯỜI GIÁM HỘ (70% có người giám hộ)
                'guardian_name': guardian_name if random.randint(1, 10) <= 7 else None,
                'guardian_job': random.choice(jobs) if random.randint(1, 10) <= 7 else None,
                'guardian_birth_year': str(random.randint(1960, 1980)) if random.randint(1, 10) <= 7 else None,
                'guardian_phone': f"0{random.randint(900000000, 999999999)}" if random.randint(1, 10) <= 7 else None,
                'guardian_cccd': f"{cccd_prefix}{random.randint(100000000, 999999999)}" if random.randint(1, 10) <= 7 else None,
                'guardian_gender': guardian_gender if random.randint(1, 10) <= 7 else None,
                
                # THÔNG TIN BỔ SUNG (80% điền)
                'organization': random.choice(['Đoàn Thanh niên', 'Đội Thiếu niên', 'Hội Học sinh']) if random.randint(1, 10) <= 8 else None,
                'job': 'Học sinh' if random.randint(1, 10) <= 9 else None,
                
                # SCHEMA COMPATIBILITY - Đảm bảo cả 2 schema đều có data
                'ho_ten': full_name,  # Schema cũ
                'ngay_sinh': birth_date,  # Schema cũ  
                'lop': class_name,  # Schema cũ
                'gioi_tinh': gender,  # Schema cũ
                'sdt': phone,  # Schema cũ
                'dan_toc': random.choice(ethnicities),  # Schema cũ
                'ton_giao': random.choice(religions),  # Schema cũ
                'tinh_thuong_tru': province,  # Schema cũ
                'phuong_xa_thuong_tru': ward,  # Schema cũ
                'cccd': cccd_number,  # Schema cũ
                'ngay_cap_cccd': cccd_date,  # Schema cũ
                'noi_cap_cccd': cccd_place,  # Schema cũ
                'ho_ten_cha': father_name,  # Schema cũ
                'sdt_cha': father_phone,  # Schema cũ
                'ho_ten_me': mother_name,  # Schema cũ
                'sdt_me': mother_phone,  # Schema cũ
            }
            
            # Filter data to only include columns that exist in database
            student_data = {}
            for key, value in all_student_data.items():
                if key in existing_columns:
                    student_data[key] = value
            
            # Đảm bảo luôn có tên học sinh trong cả 2 schema
            if 'ho_ten' in existing_columns and 'ho_ten' not in student_data:
                student_data['ho_ten'] = full_name
            if 'full_name' in existing_columns and 'full_name' not in student_data:
                student_data['full_name'] = full_name
            
            # Đảm bảo các trường quan trọng khác cũng được điền
            if 'lop' in existing_columns and 'lop' not in student_data:
                student_data['lop'] = class_name
            if 'class' in existing_columns and 'class' not in student_data:
                student_data['class'] = class_name
                
            if 'ngay_sinh' in existing_columns and 'ngay_sinh' not in student_data:
                student_data['ngay_sinh'] = birth_date
            if 'birth_date' in existing_columns and 'birth_date' not in student_data:
                student_data['birth_date'] = birth_date
                
            if 'gioi_tinh' in existing_columns and 'gioi_tinh' not in student_data:
                student_data['gioi_tinh'] = gender
            if 'gender' in existing_columns and 'gender' not in student_data:
                student_data['gender'] = gender
                
            if 'sdt' in existing_columns and 'sdt' not in student_data:
                student_data['sdt'] = phone
            if 'phone' in existing_columns and 'phone' not in student_data:
                student_data['phone'] = phone
            
            # Tính phần trăm thông tin được điền (không tính NULL)
            filled_fields = sum(1 for v in student_data.values() if v is not None and v != '' and v != 'Chưa có thông tin')
            total_fields = len(student_data)
            fill_percentage = (filled_fields / total_fields * 100) if total_fields > 0 else 0
            
            print(f"[DEBUG] Student {i+1}: {student_data.get('ho_ten') or student_data.get('full_name')} - {filled_fields}/{total_fields} fields filled ({fill_percentage:.1f}%)")
            
            # Đảm bảo đạt ít nhất 70% thông tin
            if fill_percentage < 70:
                print(f"[WARNING] Student {i+1} chỉ có {fill_percentage:.1f}% thông tin. Cần cải thiện!")
            
            print(f"[DEBUG] Using {len(student_data)} columns out of {len(all_student_data)} possible columns.")
            
            # Insert vào database
            columns = ', '.join(student_data.keys())
            placeholder = get_placeholder()
            placeholders = ', '.join([placeholder for _ in student_data.keys()])
            query = f"INSERT INTO students ({columns}) VALUES ({placeholders})"
            
            cursor.execute(query, list(student_data.values()))
            created_count += 1
        
        conn.commit()
        conn.close()
        
        print(f"[ADMIN] Đã tạo {created_count} học sinh mẫu thực tế")
        return jsonify({'success': True, 'created_count': created_count})
    except Exception as e:
        print(f"[ERROR] Generate sample data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/delete-all-students', methods=['DELETE'])
def delete_all_students():
    """Xóa tất cả học sinh"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM students")
        count = cursor.fetchone()[0]
        
        cursor.execute("DELETE FROM students")
        conn.commit()
        conn.close()
        
        print(f"[ADMIN] Đã xóa tất cả {count} học sinh")
        return jsonify({'success': True, 'deleted_count': count})
    except Exception as e:
        print(f"[ERROR] Delete all students: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generate-bots', methods=['POST'])
def generate_bots():
    """Tạo dữ liệu bot test"""
    try:
        data = request.get_json()
        count = int(data.get('count', 10))
        
        if count > 100:
            return jsonify({'success': False, 'error': 'Không thể tạo quá 100 bot cùng lúc'}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        import random
        from datetime import datetime, timedelta
        
        # Danh sách tên và lớp mẫu
        first_names = ['Nguyễn', 'Trần', 'Lê', 'Phạm', 'Hoàng', 'Huỳnh', 'Phan', 'Vũ', 'Võ', 'Đặng']
        last_names = ['Văn A', 'Thị B', 'Minh C', 'Hoàng D', 'Thùy E', 'Quang F', 'Thu G', 'Đức H', 'Mai I', 'Hải J']
        classes = [
            # Khối 10
            '10A1', '10A2', '10A3', '10A4', '10A5', '10A6', '10A7', '10A8',
            '10B1', '10B2', '10B3', '10B4',
            # Khối 11
            '11A1', '11A2', '11A3', '11A4', '11A5', '11A6', '11A7', '11A8',
            '11B1', '11B2', '11B3', '11B4',
            # Khối 12
            '12A1', '12A2', '12A3', '12A4', '12A5', '12A6', '12A7', '12A8',
            '12B1', '12B2', '12B3', '12B4'
        ]
        provinces = ['Thành phố Hồ Chí Minh', 'Thành phố Hà Nội', 'Thành phố Đà Nẵng', 'Tỉnh Bình Dương']
        wards = ['Phường Dĩ An', 'Phường Sài Gon', 'Phường Tân Bình', 'Phường Thủ Đức']
        
        created_count = 0
        for i in range(count):
            # Tạo dữ liệu ngẫu nhiên
            email = f"bot_test_{random.randint(10000, 99999)}_bot_{i}@test.com"
            full_name = f"{random.choice(first_names)} {random.choice(last_names)} (Bot)"
            class_name = random.choice(classes)
            birth_date = (datetime.now() - timedelta(days=random.randint(5840, 6570))).strftime('%Y-%m-%d')
            gender = random.choice(['Nam', 'Nữ'])
            phone = f"09{random.randint(10000000, 99999999)}"
            
            # Địa chỉ ngẫu nhiên
            province = random.choice(provinces)
            ward = random.choice(wards)
            hamlet = f"Khu phố {random.randint(1, 10)}"
            street = f"Số {random.randint(1, 999)}/{random.randint(1, 99)}, đường Test {random.randint(1, 20)}"
            
            student_data = {
                'email': email,
                'full_name': full_name,
                'class': class_name,
                'birth_date': birth_date,
                'gender': gender,
                'phone': phone,
                'permanent_province': province,
                'permanent_ward': ward,
                'permanent_hamlet': hamlet,
                'permanent_street': street,
                'hometown_province': province,
                'hometown_ward': ward,
                'hometown_hamlet': hamlet,
                'current_province': province,
                'current_ward': ward,
                'current_hamlet': hamlet,
                'current_address_detail': f"{street}, {hamlet}, {ward}, {province}",
                'birthplace_province': province,
                'birthplace_ward': ward,
                'birth_cert_province': province,
                'birth_cert_ward': ward,
                'created_at': datetime.now().isoformat()
            }
            
            # Insert vào database
            columns = ', '.join(student_data.keys())
            placeholders = ', '.join(['?' for _ in student_data.keys()])
            query = f"INSERT INTO students ({columns}) VALUES ({placeholders})"
            
            cursor.execute(query, list(student_data.values()))
            created_count += 1
        
        conn.commit()
        conn.close()
        
        print(f"[ADMIN] Đã tạo {created_count} bot test")
        return jsonify({'success': True, 'created_count': created_count})
    except Exception as e:
        print(f"[ERROR] Generate bots: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/delete-all-bots', methods=['DELETE'])
def delete_all_bots():
    """Xóa tất cả dữ liệu mẫu (ảo) - giữ lại người thật"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Đếm số data mẫu trước khi xóa (email có chứa '_sample_')
        cursor.execute("SELECT COUNT(*) FROM students WHERE email LIKE '%_sample_%'")
        count = cursor.fetchone()[0]
        
        # Xóa data mẫu (email có chứa '_sample_')
        cursor.execute("DELETE FROM students WHERE email LIKE '%_sample_%'")
        conn.commit()
        conn.close()
        
        print(f"[ADMIN] Đã xóa {count} học sinh mẫu (ảo)")
        return jsonify({'success': True, 'deleted_count': count})
    except Exception as e:
        print(f"[ERROR] Delete all sample data: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export-count', methods=['GET'])
def export_count():
    """Get count of records that would be exported with current filters"""
    try:
        # Simple test first
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Just count all records first
        cursor.execute('SELECT COUNT(*) FROM students')
        total_count = cursor.fetchone()[0]
        
        print(f"[DEBUG] Total students in database: {total_count}")
        
        # Now try with filters
        export_type = request.args.get('type', 'all')
        grade = request.args.get('grade')  
        classes = request.args.get('classes')
        province = request.args.get('province')
        ethnicity = request.args.get('ethnicity')
        gender = request.args.get('gender')
        has_phone = request.args.get('hasPhone')
        
        print(f"[DEBUG] Request params - type: {export_type}, grade: {grade}, classes: {classes}")
        print(f"[DEBUG] Filters - province: {province}, ethnicity: {ethnicity}, gender: {gender}")
        print(f"[DEBUG] Has_phone: {has_phone}")
        
        if (export_type == 'all' and not province and not ethnicity and not classes and 
            not grade and not gender and not has_phone):
            conn.close()
            return jsonify({'count': total_count})
        
        # Build filtered query
        where_conditions = []
        query_params = []
        
        if export_type == 'class' and classes:
            class_list = [cls.strip() for cls in classes.split(',')]
            placeholder = get_placeholder()
            placeholders = ','.join([placeholder for _ in class_list])
            where_conditions.append(f"class IN ({placeholders})")
            query_params.extend(class_list)
        
        if export_type == 'grade' and grade:
            placeholder = get_placeholder()
            where_conditions.append(f"class LIKE {placeholder}")
            query_params.append(f"{grade}%")
            
        if province:
            placeholder = get_placeholder()
            # Flexible province matching - case insensitive and partial match
            where_conditions.append(f"LOWER(permanent_province) LIKE LOWER({placeholder})")
            query_params.append(f"%{province}%")
            
        if ethnicity:
            placeholder = get_placeholder()
            # Flexible ethnicity matching - case insensitive and partial match
            where_conditions.append(f"LOWER(ethnicity) LIKE LOWER({placeholder})")
            query_params.append(f"%{ethnicity}%")
            
        if gender:
            gender_list = [g.strip() for g in gender.split(',')]
            placeholder = get_placeholder()
            placeholders = ','.join([placeholder for _ in gender_list])
            where_conditions.append(f"gender IN ({placeholders})")
            query_params.extend(gender_list)
            
        if has_phone and has_phone.lower() == 'true':
            where_conditions.append("sdt IS NOT NULL AND sdt != ''")
        
        if where_conditions:
            query = f"SELECT COUNT(*) FROM students WHERE {' AND '.join(where_conditions)}"
        else:
            query = "SELECT COUNT(*) FROM students"
        
        print(f"[DEBUG] Final query: {query}")
        print(f"[DEBUG] Query params: {query_params}")
        
        cursor.execute(query, query_params)
        count = cursor.fetchone()[0]
        
        print(f"[DEBUG] Filtered count: {count}")
        
        conn.close()
        return jsonify({'count': count})
        
    except Exception as e:
        print(f"[ERROR] export_count: {e}")
        # Return count 0 with success status instead of error
        return jsonify({'count': 0, 'error': str(e)}), 200

@app.route('/api/debug/provinces', methods=['GET'])
def debug_provinces():
    """Debug API để xem các tỉnh có trong database"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT DISTINCT permanent_province FROM students WHERE permanent_province IS NOT NULL ORDER BY permanent_province')
        provinces = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        return jsonify({'provinces': provinces})
        
    except Exception as e:
        print(f"[ERROR] debug_provinces: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    migrate_db()

    email_working = test_email_config()

    try:
        print("🚀 THPT Dĩ An - Server đang chạy với SSL support")
        print("🌐 HTTP: http://localhost:5000 (tự động redirect sang HTTPS)")
        print("🔒 HTTPS: https://thptdian.edu.vn (production)")
        print("👨‍💼 Admin: https://thptdian.edu.vn/admin")
        print(f"🔑 Số tài khoản admin: {len(ADMIN_ACCOUNTS)}")
        
        if SHOW_ADMIN_CREDENTIALS:
            print("🔐 Tài khoản admin:")
            for email, password in ADMIN_ACCOUNTS.items():
                print(f"   • {email} / {password}")
        else:
            print("📝 Xem file .env để biết thông tin đăng nhập")
            
        if email_working:
            print("📧 Email OTP: Đã cấu hình và hoạt động")
        else:
            print("⚠️  Email OTP: Lỗi cấu hình, sử dụng console debug")
            
        print("\n🔒 SSL Configuration:")
        print("   • Nginx reverse proxy với SSL termination")
        print("   • Let's Encrypt certificates")
        print("   • HTTP -> HTTPS redirect")
        print("   • Security headers enabled")
        print("   • Rate limiting active")
            
    except Exception:
        print("🚀 Server đang chạy với SSL support")
        print("🌐 Local: http://localhost:5000")
        print("🔒 Production: https://thptdian.edu.vn")

    port = int(os.environ.get('PORT', 5000))
    
    # Start auto-cleanup for export files
    import threading
    import glob
    
    def auto_cleanup_exports():
        """Background cleanup for export files"""
        while True:
            try:
                time.sleep(120)  # Check every 2 minutes
                patterns = ['danh_sach_*.xlsx', 'danh_sach_*.csv', 'danh_sach_*.json']
                cutoff_time = datetime.now() - timedelta(minutes=3)  # Clean files older than 3 minutes
                
                for pattern in patterns:
                    files = glob.glob(pattern)
                    for file in files:
                        try:
                            stat = os.stat(file)
                            file_time = datetime.fromtimestamp(stat.st_mtime)
                            
                            if file_time < cutoff_time:
                                os.remove(file)
                                age_minutes = (datetime.now() - file_time).total_seconds() / 60
                                print(f"🗑️ [AUTO-CLEANUP] Deleted old export: {file} (age: {age_minutes:.1f} min)")
                        except Exception as e:
                            print(f"❌ [AUTO-CLEANUP] Failed to delete {file}: {e}")
            except Exception as e:
                print(f"❌ [AUTO-CLEANUP] Error in cleanup loop: {e}")
                time.sleep(60)
    
    # Start cleanup thread
    cleanup_thread = threading.Thread(target=auto_cleanup_exports)
    cleanup_thread.daemon = True
    cleanup_thread.start()
    print("🧹 Auto-cleanup started for export files")
    
    app.run(debug=False, host='0.0.0.0', port=port)
